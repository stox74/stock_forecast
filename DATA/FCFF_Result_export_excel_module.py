"""
FCFF_Result_export_excel_module.py
==================================
Korea FCFF DCF v2 — Excel Export Module

미국 v11 모형의 export_v10_excel 기능을 한국 데이터스키마에 맞게 한국화하여
별도 모듈로 분리. 9-sheet workbook 빌드 (Main / Verification / Phase2 / Moat /
Beta / CoC / DCF / Sensitivity / Notes).

이 모듈을 두 PC 모두의 프로젝트 루트 (Cell 1 의 `_ROOT`) 직속에 배치하면
노트북에서 `import FCFF_Result_export_excel_module` 한 줄로 사용 가능.

------------------------------------------------------------------------
사용법
------------------------------------------------------------------------
    # (1) 노트북에서:
    from FCFF_Result_export_excel_module import (
        configure, register_excel_export
    )

    # (2) 한국 모형의 상수·로거 전달 (선택 — default 값 그대로 써도 됨):
    configure(
        debt_keys=DEBT_KEYS,                                # Cell 2
        cash_keys=CASH_KEYS,
        log_func=log,                                       # 한국 모형 내부 log
        marketcap_table="ks_listed_company_daily_marketcap",
    )

    # (3) 클래스에 메서드 attach (1회):
    register_excel_export(KoreaDCFModel)

    # (4) 사용:
    model = KoreaDCFModel(...).run()
    out_path = model.export_korea_excel(out_dir="C:/reports")

------------------------------------------------------------------------
미국 vs 한국 핵심 매핑
------------------------------------------------------------------------
    self._inc/_cf/_bs       → self._fs_wide   (단일 wide-form)
    inc["operatingIncome"]  → wide["operating_income"]
    inc["revenue"]          → wide["revenue"]
    inc["interestExpense"]  → wide["interest_expense"]
    cf["depreciationAnd…"]  → wide["da_cf"] + wide["intangible_amort_cf"]
    cf["capitalExpenditure"]→ wide["capex_tangible"] + wide["capex_intangible"]
    bs["totalDebt"]         → Σ(DEBT_KEYS)
    bs["cashAndCashEq…"]    → Σ(CASH_KEYS)
    self.api_key (FMP)      → 사용 안함 (DB 조회로 대체)
    self.ticker (MSFT)      → self.ticker_dg (A005930)
    ^GSPC (S&P500)          → KOSPI
    _RF/_RM_DEFAULT         → self.rf / self.e_rm
    self._profile           → self.valuation 에서 추출
    $ ($#,##0.00)           → 원 (#,##0"원")
"""

# 모듈 버전 (옵션 B 적용 = v2.2, Main 시트 정합성 = v2.3)
# 이 문자열이 'v2.3-mainsheet-fix' 이 아니면 옛 버전이 import 된 것이므로
# DATA/ 폴더에 새 파일을 덮어쓰지 않았거나, Jupyter 가 옛 import 를 캐시하고 있음
__version__ = "v2.3-mainsheet-fix"

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path


# ════════════════════════════════════════════════════════════════════
#  모듈 설정 — configure() 로 override 가능
# ════════════════════════════════════════════════════════════════════

_CONFIG = {
    # NWC / 자본구조 합산용 키 (한국 v2 02_imports.py 와 동일 default)
    "DEBT_KEYS": ["short_term_debt", "current_lt_debt", "bonds",
                  "long_term_debt", "lease_liab"],
    "CASH_KEYS": ["cash", "short_term_invest"],
    # 외부 logger (None 이면 print 사용)
    "log_func": None,
    # 일별 시세 DB 테이블
    "marketcap_table": "ks_listed_company_daily_marketcap",
    # Blume 가중치 (호영님 기준 0.67/0.33, 그대로 둠)
    "BLUME_W_RAW": 0.67,
    "BLUME_W_MARKET": 0.33,
}


def configure(debt_keys=None, cash_keys=None, log_func=None,
              marketcap_table=None,
              blume_w_raw=None, blume_w_market=None):
    """
    모듈 설정 override. register_excel_export() 호출 전에 사용.

    Parameters
    ----------
    debt_keys : list[str], optional
        총부채 합산용 BS 컬럼 키 (default: 한국 v2 표준 5개).
    cash_keys : list[str], optional
        현금성 자산 합산용 BS 컬럼 키 (default: 2개).
    log_func : callable, optional
        외부 로거 함수. signature: log_func(ticker, msg).
        None 이면 print() 사용.
    marketcap_table : str, optional
        일별 시세 DB 테이블명. default 'ks_listed_company_daily_marketcap'.
    blume_w_raw, blume_w_market : float, optional
        Blume 보정 가중치 (default: 0.67, 0.33).
    """
    global _CONFIG
    if debt_keys       is not None: _CONFIG["DEBT_KEYS"]       = debt_keys
    if cash_keys       is not None: _CONFIG["CASH_KEYS"]       = cash_keys
    if log_func        is not None: _CONFIG["log_func"]        = log_func
    if marketcap_table is not None: _CONFIG["marketcap_table"] = marketcap_table
    if blume_w_raw     is not None: _CONFIG["BLUME_W_RAW"]     = blume_w_raw
    if blume_w_market  is not None: _CONFIG["BLUME_W_MARKET"]  = blume_w_market


def _log(ticker, msg):
    """모듈 내부 logger — 외부 log_func 우선, 없으면 print."""
    fn = _CONFIG.get("log_func")
    if fn is not None:
        fn(ticker, msg)
    else:
        print(f"[{ticker}] {msg}")


# ════════════════════════════════════════════════════════════════════
#  스타일 상수 (미국 v10 LD1 동일)
# ════════════════════════════════════════════════════════════════════

# ── 색상 팔레트 ─────────────────────────────────────────────────
_NAVY, _TEAL, _GOLD       = "0D1B2A", "0E7C7B", "F4A261"
_ZEBRA, _FCST_BG, _ASSUMP = "F5F7FA", "FFF4D6", "FFF9C4"
_GREY, _BLUE              = "D9D9D9", "0000FF"
_PHASE1_BG, _PHASE2_BG, _PHASE3_BG = "E8F4F8", "FFF4D6", "F0E5C8"
_LINK_BG = "E0F0FF"

# ── 숫자 포맷 (★ 한국화: $ → 원) ────────────────────────────────
_FMT_INT, _FMT_PCT, _FMT_PCT3, _FMT_DEC = (
    "#,##0;[Red](#,##0);-", "0.00%", "0.000%", "0.0000"
)
_FMT_PRICE = '#,##0"원"'   # 미국 "$#,##0.00" → 한국 원 표기

_THIN   = Side(style="thin", color=_GREY)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ════════════════════════════════════════════════════════════════════
#  Style helpers
# ════════════════════════════════════════════════════════════════════

def _set_label(cell, text, bold=False, italic=False, size=10, color="000000"):
    cell.value = text
    cell.font  = Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)
    cell.alignment = Alignment(horizontal="left")
    cell.border = _BORDER


def _set_input(cell, value, fmt):
    cell.value = value
    cell.number_format = fmt
    cell.fill = PatternFill("solid", start_color=_ASSUMP)
    cell.font = Font(name="Calibri", size=10, color=_BLUE, bold=True)
    cell.alignment = Alignment(horizontal="right")
    cell.border = _BORDER


def _set_computed(cell, value, fmt, highlight=False, final=False):
    cell.value = value
    cell.number_format = fmt
    if final:
        cell.fill = PatternFill("solid", start_color=_NAVY)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    elif highlight:
        cell.fill = PatternFill("solid", start_color=_GOLD)
        cell.font = Font(name="Calibri", size=11, bold=True)
    else:
        cell.fill = PatternFill("solid", start_color=_LINK_BG)
        cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="right")
    cell.border = _BORDER


def _section_header(ws, row, text, span=6, color=_TEAL):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    h = ws.cell(row=row, column=1, value=text)
    h.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", start_color=color)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


# ════════════════════════════════════════════════════════════════════
#  ① _fetch_korea_extras  (10Y monthly close: 종목 + KOSPI)
# ════════════════════════════════════════════════════════════════════

def _fetch_korea_extras(self, years: int = 10) -> None:
    """
    Excel export 에 필요한 추가 데이터 fetch.
    self._monthly_prices, self._monthly_returns 에 저장.

    한국 데이터 출처:
      - 종목 일별 종가: _CONFIG["marketcap_table"] (default: ks_listed_company_daily_marketcap)
      - KOSPI 일별 종가: self.kospi (이미 인스턴스에 있음)
    """
    import pymysql

    end_dt   = datetime.today()
    start_dt = end_dt.replace(year=end_dt.year - years)
    table    = _CONFIG["marketcap_table"]

    # ── 종목 일별 → 월말 리샘플 ──────────────────────────────
    ticker_monthly = pd.Series(dtype=float)
    try:
        conn = pymysql.connect(**self.db_info)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    f"SELECT date, close_price FROM {table} "
                    "WHERE ticker = %s AND date >= %s AND close_price > 0 "
                    "ORDER BY date",
                    (self.ticker_dg,
                     start_dt.strftime("%Y-%m-%d"))
                )
                rows = cur.fetchall()
        finally:
            conn.close()

        if rows:
            df = pd.DataFrame(rows)
            df["date"] = pd.to_datetime(df["date"])
            df = df.set_index("date").sort_index()
            ticker_monthly = (df["close_price"]
                              .resample("ME").last().dropna()
                              .tail(years * 12))
    except Exception as e:
        if getattr(self, "verbose", False):
            _log(self.ticker_dg, f"종목 월별 종가 fetch fail: {e}")

    # ── KOSPI 월말 리샘플 ───────────────────────────────────
    kospi_monthly = pd.Series(dtype=float)
    if getattr(self, "kospi", None) is not None and not self.kospi.empty:
        try:
            kospi_idx = pd.to_datetime(self.kospi.index)
            kospi_ser = pd.Series(self.kospi.values, index=kospi_idx)
            kospi_monthly = (kospi_ser
                             .resample("ME").last().dropna()
                             .tail(years * 12))
        except Exception as e:
            if getattr(self, "verbose", False):
                _log(self.ticker_dg, f"KOSPI 월별 리샘플 fail: {e}")

    # ── 결합 ────────────────────────────────────────────────
    if not ticker_monthly.empty and not kospi_monthly.empty:
        self._monthly_prices = pd.DataFrame({
            self.ticker_dg: ticker_monthly,
            "KOSPI":        kospi_monthly,
        }).dropna()
        self._monthly_returns = self._monthly_prices.pct_change().dropna()
    else:
        self._monthly_prices  = pd.DataFrame()
        self._monthly_returns = pd.DataFrame()

    if getattr(self, "verbose", False):
        _log(self.ticker_dg,
             f"Korea extras: {len(self._monthly_returns)}개월 returns "
             f"(종목 {len(ticker_monthly)}, KOSPI {len(kospi_monthly)})")


# ════════════════════════════════════════════════════════════════════
#  과거 FCFF 구성요소 (Sales/OPM/EBIT/.../FCFF) — 40Q rebuild
# ════════════════════════════════════════════════════════════════════

def _build_hist_components_kr(model, n_qtrs: int = 40) -> pd.DataFrame:
    """
    KoreaDCFModel._compute_historical_fcff 와 동일 로직이지만 모든 구성요소
    (Sales, OPM, EBIT, NOPAT, DA, CapEx, NWC, ΔNWC, FCFF) 캡처.
    Excel main sheet 채우기용.

    ★ v8 패치: 첫 분기 prev_nwc = γ × first_sales 시드값 적용
    """
    sales_act = model._sales_actual
    if sales_act is None or sales_act.empty:
        return pd.DataFrame()

    tax = model.estimate_tax_rate()
    alpha, _ = model.estimate_ratio_coef(
        ["da_cf", "intangible_amort_cf"], "da")
    beta_, _ = model.estimate_ratio_coef(
        ["capex_tangible", "capex_intangible"], "capex", take_abs=True)
    gamma, _ = model.estimate_nwc_coef()

    wide = model._fs_wide
    rows = []
    first_sales = float(sales_act.iloc[0])
    prev_nwc = gamma * first_sales   # ★ v8 시드

    for dt, sales in sales_act.items():
        if pd.isna(sales) or sales <= 0:
            prev_nwc = gamma * sales if pd.notna(sales) else prev_nwc
            continue
        if dt not in wide.index:
            prev_nwc = gamma * sales
            continue
        op_v  = wide.loc[dt].get("operating_income", np.nan)
        rev_v = wide.loc[dt].get("revenue", np.nan)
        opm = (float(op_v) / float(rev_v)
               if (pd.notna(op_v) and pd.notna(rev_v) and float(rev_v) > 0)
               else np.nan)
        if pd.isna(opm):
            prev_nwc = gamma * sales
            continue
        ebit  = sales * opm
        nopat = ebit * (1 - tax)
        da    = alpha * sales
        capex = beta_ * sales
        nwc   = gamma * sales
        dnwc  = nwc - prev_nwc
        prev_nwc = nwc
        fcff  = nopat + da - capex - dnwc
        rows.append({
            "date": dt,
            "quarter": f"{dt.year}Q{dt.quarter}",
            "Sales": sales, "OPM": opm, "EBIT": ebit, "NOPAT": nopat,
            "DA": da, "CapEx": capex, "NWC": nwc, "ΔNWC": dnwc, "FCFF": fcff,
        })
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).tail(n_qtrs).reset_index(drop=True)
def _build_main_kr(model, wb, hist_df, fc_df, ticker_dg):
    ws = wb.active
    ws.title = f"{ticker_dg}_FCFF"

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"{ticker_dg} — 분기별 FCFF 구성요소 + 8Q 예측 (단위: 원)"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["분기","Sales","OPM","EBIT","NOPAT","D&A","CapEx","NWC","ΔNWC","FCFF","연간 FCFF (4Q sum)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_TEAL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
    ws.row_dimensions[2].height = 24

    N_HIST = len(hist_df)

    # ── 과거 분기 (rows 3..2+N_HIST) ─────────────────────────────
    for i in range(N_HIST):
        rd = hist_df.iloc[i]
        r = i + 3
        ws.cell(row=r, column=1, value=rd["quarter"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=float(rd["Sales"])).number_format = _FMT_INT
        ws.cell(row=r, column=3, value=float(rd["OPM"])).number_format  = _FMT_PCT
        ws.cell(row=r, column=4, value=float(rd["EBIT"])).number_format = _FMT_INT
        ws.cell(row=r, column=5, value=float(rd["NOPAT"])).number_format= _FMT_INT
        ws.cell(row=r, column=6, value=float(rd["DA"])).number_format   = _FMT_INT
        ws.cell(row=r, column=7, value=float(rd["CapEx"])).number_format= _FMT_INT
        ws.cell(row=r, column=8, value=float(rd["NWC"])).number_format  = _FMT_INT
        if i == 0:
            for c_idx in (9, 10):
                cc = ws.cell(row=r, column=c_idx, value="–")
                cc.alignment = Alignment(horizontal="center")
        else:
            ws.cell(row=r, column=9,  value=f"=H{r}-H{r-1}").number_format = _FMT_INT
            ws.cell(row=r, column=10, value=f"=E{r}+F{r}-G{r}-I{r}").number_format = _FMT_INT
        # 연간 합계 (K열) — 매 4Q (Q4) 끝에만 SUM(J)
        try:
            qnum = int(str(rd["quarter"])[-1])
        except Exception:
            qnum = 0
        if qnum == 4 and i >= 3:
            ws.cell(row=r, column=11, value=f"=SUM(J{r-3}:J{r})").number_format = _FMT_INT
        fill = PatternFill("solid", start_color=_ZEBRA) if i % 2 == 0 else None
        for col in range(1, 12):
            cc = ws.cell(row=r, column=col)
            cc.border = _BORDER
            cc.font = Font(name="Calibri", size=10)
            if fill: cc.fill = fill

    # 향후 8Q (수식 — 가정셀 참조)
    # ── 향후 8Q forecast ─────────────────────────────────────────
    # ★ v2.3 변경: Excel 수식 (평균 OPM 적용) → model.result_df 직접 입력
    #   이로써 Main 시트의 4Q 합계 (K열) 가 Phase2_LongTerm 의 ph1_annual
    #   (Year 1, Year 2 FCFF) 와 자동 일치.
    fcst_top  = Side(style="medium", color=_GOLD)
    fcst_fill = PatternFill("solid", start_color=_FCST_BG)
    fcst_start_row = 3 + N_HIST

    for j in range(min(8, len(fc_df))):
        rd = fc_df.iloc[j]
        fc_label = rd.get("quarter", f"F{j+1}")
        r = fcst_start_row + j

        # ★ 모든 값 model.result_df 에서 직접 가져옴 (수식 아님)
        ws.cell(row=r, column=1, value=str(fc_label)).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=float(rd.get("sales_forecast", np.nan))).number_format = _FMT_INT
        ws.cell(row=r, column=3, value=float(rd.get("opm_forecast",   np.nan))).number_format = _FMT_PCT
        ws.cell(row=r, column=4, value=float(rd.get("ebit",  np.nan))).number_format = _FMT_INT
        ws.cell(row=r, column=5, value=float(rd.get("nopat", np.nan))).number_format = _FMT_INT
        ws.cell(row=r, column=6, value=float(rd.get("da",    np.nan))).number_format = _FMT_INT
        ws.cell(row=r, column=7, value=float(rd.get("capex", np.nan))).number_format = _FMT_INT
        ws.cell(row=r, column=8, value=float(rd.get("nwc",   np.nan))).number_format = _FMT_INT
        ws.cell(row=r, column=9, value=float(rd.get("delta_nwc", np.nan))).number_format = _FMT_INT
        ws.cell(row=r, column=10, value=float(rd.get("fcff",  np.nan))).number_format = _FMT_INT

        # 연간 합계 (K열) — 매 4Q (Q4) 끝에만 SUM(J)
        try:
            qnum = int(str(fc_label)[-1])
        except Exception:
            qnum = 0
        if qnum == 4 and j >= 3:
            ws.cell(row=r, column=11, value=f"=SUM(J{r-3}:J{r})").number_format = _FMT_INT

        for col in range(1, 12):
            cc = ws.cell(row=r, column=col)
            cc.border = Border(left=_THIN, right=_THIN, bottom=_THIN,
                               top=fcst_top if j == 0 else _THIN)
            cc.fill = fcst_fill
            cc.font = Font(name="Calibri", size=10)

    # ── 참고용: model 의 forecast 가정값 평균 (★ v2.3: 더 이상 수식 입력 안됨) ────
    def _quarterly_yoy(df):
        g = {1: np.nan, 2: np.nan, 3: np.nan, 4: np.nan}
        df = df.copy()
        df["q_num"] = df["quarter"].astype(str).str[-1].astype(int)
        for q in [1, 2, 3, 4]:
            sq = df.loc[df["q_num"] == q].sort_values("quarter")["Sales"].values
            if len(sq) >= 2:
                yoy = sq[1:] / sq[:-1] - 1
                g[q] = float(np.prod(1 + yoy)) ** (1.0 / len(yoy)) - 1
        return g

    growth = _quarterly_yoy(hist_df)

    # M열 가정값 — ★ v2.3: 참고용으로 변경 (편집해도 forecast 영역에 영향 없음)
    _set_label(ws.cell(row=10, column=12), "Forecast 가정값 (참고용 — model 평균)", bold=True)
    ws.cell(row=10, column=12).fill = PatternFill("solid", start_color=_TEAL)
    ws.cell(row=10, column=12).font = Font(bold=True, color="FFFFFF")

    fc_assumptions = {
        "OPM (mean)":          float(fc_df["opm_forecast"].mean()) if "opm_forecast" in fc_df.columns else 0.10,
        "Tax rate":            float(fc_df["tax_rate"].iloc[0])    if "tax_rate"     in fc_df.columns else 0.21,
        "D&A / Sales (mean)":  float(fc_df["da"].mean()    / fc_df["sales_forecast"].mean()) if "sales_forecast" in fc_df.columns and fc_df["sales_forecast"].mean() > 0 else 0.05,
        "CapEx / Sales (mean)":float(fc_df["capex"].mean() / fc_df["sales_forecast"].mean()) if "sales_forecast" in fc_df.columns and fc_df["sales_forecast"].mean() > 0 else 0.07,
        "NWC / Sales (mean)":  float(fc_df["nwc"].mean()   / fc_df["sales_forecast"].mean()) if "sales_forecast" in fc_df.columns and fc_df["sales_forecast"].mean() > 0 else 0.15,
    }
    for i, (k, v) in enumerate(fc_assumptions.items(), start=11):
        _set_label(ws.cell(row=i, column=12), k)
        # ★ v2.3: 노란 input 이 아니라 참고용 회색 표시 (수식 참조 없음)
        c = ws.cell(row=i, column=13, value=v)
        c.number_format = _FMT_PCT
        c.fill = PatternFill("solid", start_color=_LINK_BG)
        c.font = Font(name="Calibri", size=10, italic=True, color="595959")
        c.alignment = Alignment(horizontal="right")
        c.border = _BORDER

    # 안내 메모
    ws.merge_cells("L17:N17")
    note_ref = ws.cell(row=17, column=12,
                       value="  ※ 참고용 — Forecast 영역(R43+) 은 model.result_df 직접 입력")
    note_ref.font = Font(italic=True, size=9, color="595959")
    note_ref.alignment = Alignment(horizontal="left")

    # N열 분기별 성장률 (역사적 추정값 — 참고용)
    _set_label(ws.cell(row=2, column=14), "분기별 YoY 성장률 (참고용 — 과거 hist)", bold=True)
    ws.cell(row=2, column=14).fill = PatternFill("solid", start_color=_TEAL)
    ws.cell(row=2, column=14).font = Font(bold=True, color="FFFFFF")
    ws.merge_cells("N2:N3")
    for q in [1, 2, 3, 4]:
        _set_label(ws.cell(row=3 + q, column=13), f"Q{q}")
        # ★ v2.3: 노란 input 이 아닌 참고용 회색 (forecast 영역과 무관)
        gv = growth[q] if not np.isnan(growth[q]) else 0.0
        c = ws.cell(row=3 + q, column=14, value=gv)
        c.number_format = _FMT_PCT3
        c.fill = PatternFill("solid", start_color=_LINK_BG)
        c.font = Font(name="Calibri", size=10, italic=True, color="595959")
        c.alignment = Alignment(horizontal="right")
        c.border = _BORDER

    # 컬럼 폭
    ws.column_dimensions["A"].width = 12
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["K"].width = 22  # ★ v2.3: 연간 합계 (4Q sum) 컬럼 추가
    ws.column_dimensions["L"].width = 22
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 14
    ws.freeze_panes = "B3"


# ════════════════════════════════════════════════════════════════════
#  Sheet 2 · Verification  (YoY geo mean by quarter — cross-check)
# ════════════════════════════════════════════════════════════════════

def _build_verification_kr(wb, hist_df, ticker_dg):
    ws = wb.create_sheet("Verification")
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"{ticker_dg} — YoY 성장률 교차검증 (분기별 기하평균)"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    headers = ["분기 번호", "관측 수", "기하평균 YoY", "산술평균 YoY"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_TEAL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    df = hist_df.copy()
    df["q_num"] = df["quarter"].astype(str).str[-1].astype(int)
    for q in [1, 2, 3, 4]:
        sq = df.loc[df["q_num"] == q].sort_values("quarter")["Sales"].values
        r = 3 + q
        if len(sq) >= 2:
            yoy = sq[1:] / sq[:-1] - 1
            arith = float(np.mean(yoy))
            geo   = float(np.prod(1 + yoy)) ** (1.0 / len(yoy)) - 1
            n = len(yoy)
        else:
            arith, geo, n = np.nan, np.nan, 0
        _set_label(ws.cell(row=r, column=1), f"Q{q}")
        ws.cell(row=r, column=2, value=n).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).border = _BORDER
        if n > 0:
            _set_computed(ws.cell(row=r, column=3), geo,   _FMT_PCT3)
            _set_computed(ws.cell(row=r, column=4), arith, _FMT_PCT3)
        else:
            ws.cell(row=r, column=3, value="N/A").alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=4, value="N/A").alignment = Alignment(horizontal="center")

    for col in "ABCD":
        ws.column_dimensions[col].width = 22


# ════════════════════════════════════════════════════════════════════
#  Sheet 3 · Phase2_LongTerm  (Year 1-22 AR(1) decay → g_term)
# ════════════════════════════════════════════════════════════════════

def _build_phase2_kr(wb, model, ticker_dg):
    ws = wb.create_sheet("Phase2_LongTerm")

    val = model.valuation
    ph1 = val.get("ph1_annual", [])
    ph2 = val.get("ph2_annual", [])
    g_path = val.get("ph2_growth", [])
    rho     = val.get("moat_rho", 0.83)
    g_term  = val.get("g_terminal", 0.04)
    g0      = val.get("g0", np.nan)
    g0_method = val.get("g0_method", "n/a")
    moat = val.get("moat_label", "N/A")
    n_ph2 = len(ph2)

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"{ticker_dg} — Phase 2 장기 FCFF (AR(1) decay → g_term, 단위: 원)"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Inputs
    _section_header(ws, 3, "Inputs (편집 가능)", 6, _TEAL)
    _set_label(ws.cell(row=4, column=1), "ρ (AR(1) persistence)")
    _set_input(ws.cell(row=4, column=2), float(rho), _FMT_DEC)
    _set_label(ws.cell(row=5, column=1), "g₀ (initial growth)")
    _set_input(ws.cell(row=5, column=2),
               float(g0) if (g0 is not None and not (isinstance(g0, float) and np.isnan(g0))) else 0.0,
               _FMT_PCT)
    _set_label(ws.cell(row=6, column=1), "g_term (terminal growth)")
    _set_input(ws.cell(row=6, column=2), float(g_term), _FMT_PCT)
    _set_label(ws.cell(row=7, column=1), "Phase 2 horizon (years)")
    _set_input(ws.cell(row=7, column=2), int(n_ph2), "0")

    _set_label(ws.cell(row=9, column=1),
               f"Moat:  {moat}    g₀ method:  {g0_method}",
               italic=True, color="595959")

    # Year-by-Year table
    _section_header(ws, 12, "연도별 FCFF Schedule (Year 1-22)", 6, _TEAL)
    for c, h in enumerate(["Year","Cal. Year","Phase","g(t)","FCFF","Notes"], 1):
        cell = ws.cell(row=13, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    # 시작 연도 — 예측 첫 분기에서 추출
    try:
        BASE_YEAR = int(str(model.result_df["quarter"].iloc[0])[:4])
    except Exception:
        BASE_YEAR = datetime.now().year

    ph1_growth = []
    if len(ph1) >= 2 and ph1[0] != 0:
        ph1_growth = [np.nan, ph1[1]/ph1[0] - 1]
    elif len(ph1) >= 1:
        ph1_growth = [np.nan]

    # Year 1-2: Phase 1 (DB forecast)
    for y in range(min(2, len(ph1))):
        r = 14 + y
        ws.cell(row=r, column=1, value=y+1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=BASE_YEAR + y).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value="Phase 1").alignment = Alignment(horizontal="center")
        if y < len(ph1_growth) and not (isinstance(ph1_growth[y], float) and np.isnan(ph1_growth[y])):
            _set_computed(ws.cell(row=r, column=4), float(ph1_growth[y]), _FMT_PCT)
        else:
            ws.cell(row=r, column=4, value="—").alignment = Alignment(horizontal="center")
        _set_computed(ws.cell(row=r, column=5), float(ph1[y]), _FMT_INT)
        ws.cell(row=r, column=6, value="Forecast 8Q sum").font = Font(italic=True, size=9, color="595959")
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=_PHASE1_BG)
            ws.cell(row=r, column=col).border = _BORDER

    # Year 3 ~ Year (2+n_ph2): Phase 2 (AR(1))
    last_fcff = float(ph1[-1]) if ph1 else 1.0
    for j in range(n_ph2):
        y_idx = 2 + j
        r = 14 + y_idx
        ws.cell(row=r, column=1, value=y_idx+1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=BASE_YEAR + y_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value="Phase 2").alignment = Alignment(horizontal="center")
        _set_computed(ws.cell(row=r, column=4),
                      float(g_path[j]) if j < len(g_path) else np.nan, _FMT_PCT)
        _set_computed(ws.cell(row=r, column=5), float(ph2[j]),     _FMT_INT)
        ws.cell(row=r, column=6, value=f"AR(1) decay  ρ={rho:.3f}").font = Font(italic=True, size=9, color="595959")
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=_PHASE2_BG)
            ws.cell(row=r, column=col).border = _BORDER
        last_fcff = float(ph2[j])

    # Phase 3 padding (Year n+1 ~ Year 22) — TV growth 적용
    last_year_done = 2 + n_ph2
    for y_idx in range(last_year_done, 22):
        r = 14 + y_idx
        last_fcff = last_fcff * (1 + g_term)
        ws.cell(row=r, column=1, value=y_idx+1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=BASE_YEAR + y_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value="Phase 3").alignment = Alignment(horizontal="center")
        _set_computed(ws.cell(row=r, column=4), float(g_term), _FMT_PCT)
        _set_computed(ws.cell(row=r, column=5), last_fcff,     _FMT_INT)
        ws.cell(row=r, column=6, value="Terminal growth").font = Font(italic=True, size=9, color="595959")
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=_PHASE3_BG)
            ws.cell(row=r, column=col).border = _BORDER

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28


# ════════════════════════════════════════════════════════════════════
#  Sheet 4 · Moat_Comparison  (4-tier ρ reference)
# ════════════════════════════════════════════════════════════════════

def _build_moat_kr(wb, model, ticker_dg):
    ws = wb.create_sheet("Moat_Comparison")
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"{ticker_dg} — Moat 분류 & ρ 비교"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    _section_header(ws, 3, "EVA Spread (현재 시점)", 8, _TEAL)
    val = model.valuation
    eva_cache = getattr(model, "_eva_cache", {}) or {}
    rows_eva = [
        ("ROIC (TTM)",                  val.get("roic", np.nan),       _FMT_PCT),
        ("WACC",                        val.get("wacc", np.nan),       _FMT_PCT),
        ("EVA Spread = ROIC − WACC",    val.get("eva_spread", np.nan), _FMT_PCT),
        ("Positive quarters / 20Q",     eva_cache.get("n_positive", 0),"0"),
        ("현재 Moat 등급",              val.get("moat_label", "N/A"),  None),
        ("ρ (AR(1) persistence) — selected", val.get("moat_rho", np.nan), _FMT_DEC),
        ("Phase 2 horizon (years)",     val.get("n_phase2", 0),        "0"),
    ]
    for i, (label, v, fmt) in enumerate(rows_eva):
        r = 5 + i
        _set_label(ws.cell(row=r, column=1), label)
        if isinstance(v, str):
            ws.cell(row=r, column=2, value=v).alignment = Alignment(horizontal="right")
        elif fmt is None:
            ws.cell(row=r, column=2, value=v).alignment = Alignment(horizontal="right")
        else:
            cv = ws.cell(row=r, column=2, value=v if (v is not None and not (isinstance(v, float) and np.isnan(v))) else "N/A")
            if not isinstance(cv.value, str):
                cv.number_format = fmt
                cv.fill = PatternFill("solid", start_color=_LINK_BG)
            cv.alignment = Alignment(horizontal="right")
            cv.border = _BORDER

    _section_header(ws, 14, "Moat 등급 참고표 (Damodaran / Mauboussin)", 8, _GOLD)
    headers = ["Moat", "EVA Spread Threshold", "n_positive Threshold",
               "ρ (decay)", "Phase 2 Years", "Approx. half-life (yrs)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=16, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    tiers = [
        ("Wide moat",   "> 15%",  "≥ 15", 0.90, 15),
        ("Narrow moat", "> 8%",   "≥ 12", 0.83, 12),
        ("Some moat",   "> 3%",   "≥ 8",  0.75,  8),
        ("No moat",     "≤ 3%",   "< 8",  0.60,  5),
    ]
    cur_moat = val.get("moat_label", "N/A")
    for i, (lbl, sp, npos, rho_v, yr) in enumerate(tiers):
        r = 17 + i
        is_current = (lbl == cur_moat)
        bg = _GOLD if is_current else _ZEBRA
        for col, v in enumerate([lbl, sp, npos, rho_v, yr,
                                 round(np.log(0.5)/np.log(rho_v), 1)], 1):
            cc = ws.cell(row=r, column=col, value=v)
            cc.font = Font(bold=is_current)
            cc.fill = PatternFill("solid", start_color=bg)
            cc.alignment = Alignment(horizontal="center")
            cc.border = _BORDER
            if isinstance(v, float):
                cc.number_format = _FMT_DEC

    ws.merge_cells("A22:H22")
    ws.cell(row=22, column=1,
            value="  ★ 굵은 행이 현재 종목의 등급. ρ → Phase 2 기간 매핑은 EVA spread 의 지속성에 비례.").font = Font(italic=True, size=9, color="595959")

    ws.column_dimensions["A"].width = 36
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18


# (계속 — 나머지 5개 sheet builder 와 export_korea_excel main 은 후속 패치에서 정의)


# ════════════════════════════════════════════════════════════════════
#  Sheet 5 · Beta_Regression  (10Y monthly returns vs KOSPI)
# ════════════════════════════════════════════════════════════════════

def _build_beta_kr(wb, model, ticker_dg):
    """10Y monthly returns + SLOPE/INTERCEPT/RSQ + Blume."""
    br = wb.create_sheet("Beta_Regression")
    br.merge_cells("A1:F1")
    t = br["A1"]
    t.value = f"{ticker_dg} — Beta 추정: 10년 월간 수익률 vs KOSPI"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    br.row_dimensions[1].height = 32

    for c, h in enumerate(["Date", f"{ticker_dg} Price", "KOSPI",
                            f"{ticker_dg} Return", "KOSPI Return"], 1):
        cell = br.cell(row=3, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    prices = getattr(model, "_monthly_prices", pd.DataFrame())
    if prices is None or prices.empty or len(prices) < 12:
        br.cell(row=4, column=1,
                value="(월별 종가 데이터 부족 — _fetch_korea_extras 결과 없음)").font = Font(italic=True, color="595959")
        br.column_dimensions["A"].width = 36
        for col in "BCDE": br.column_dimensions[col].width = 14
        return None, None

    N = len(prices)
    for i, (date, row) in enumerate(prices.iterrows()):
        r = 4 + i
        br.cell(row=r, column=1, value=date.date()).number_format = "yyyy-mm"
        br.cell(row=r, column=2, value=float(row[ticker_dg])).number_format = "#,##0"
        br.cell(row=r, column=3, value=float(row["KOSPI"])).number_format = "0.00"
        if i > 0:
            br.cell(row=r, column=4, value=f"=B{r}/B{r-1}-1").number_format = _FMT_PCT3
            br.cell(row=r, column=5, value=f"=C{r}/C{r-1}-1").number_format = _FMT_PCT3
        for c in range(1, 6):
            br.cell(row=r, column=c).border = _BORDER

    RET_START = 5
    RET_END   = 4 + N - 1
    RES_HDR   = N + 5
    BETA_R    = RES_HDR + 1
    ALPHA_R   = RES_HDR + 2
    BLUME_R   = RES_HDR + 14

    _section_header(br, RES_HDR, "Regression Statistics — OLS Beta Estimation", 5, _TEAL)
    for offset, label, formula, fmt, hl in [
        (1, "Beta (slope)",                       f"=SLOPE(D{RET_START}:D{RET_END},E{RET_START}:E{RET_END})",     "0.0000", True),
        (2, "Intercept (alpha, monthly)",         f"=INTERCEPT(D{RET_START}:D{RET_END},E{RET_START}:E{RET_END})", "0.000%", False),
        (3, "Annualized alpha (×12)",             f"=B{ALPHA_R}*12",                                              "0.00%",  False),
        (4, "R² (coefficient of determination)",  f"=RSQ(D{RET_START}:D{RET_END},E{RET_START}:E{RET_END})",       "0.000",  False),
        (5, "Number of observations",             f"=COUNT(D{RET_START}:D{RET_END})",                             "0",      False),
    ]:
        c1 = br.cell(row=RES_HDR + offset, column=1, value=label)
        c1.font = Font(name="Calibri", size=10, bold=hl)
        c1.alignment = Alignment(horizontal="left")
        c1.border = _BORDER
        c2 = br.cell(row=RES_HDR + offset, column=2, value=formula)
        _set_computed(c2, formula, fmt, highlight=hl)

    _section_header(br, RES_HDR + 7, "Blume 보정 — Beta 평균회귀 (호영님 기준 0.67/0.33)", 5, _TEAL)
    br.merge_cells(f"A{RES_HDR + 8}:E{RES_HDR + 8}")
    note = br.cell(row=RES_HDR + 8, column=1,
                   value="  Beta_Blume = w_raw × Beta_raw + w_market × 1.0  (한국 표준: 2/3, 1/3)")
    note.font = Font(italic=True, size=10, color="595959")
    _set_label(br.cell(row=RES_HDR + 10, column=1), "Weight on raw Beta (w_raw)")
    _set_input(br.cell(row=RES_HDR + 10, column=2), _CONFIG["BLUME_W_RAW"], _FMT_DEC)
    _set_label(br.cell(row=RES_HDR + 11, column=1), "Weight on market Beta = 1 (w_market)")
    _set_input(br.cell(row=RES_HDR + 11, column=2), _CONFIG["BLUME_W_MARKET"], _FMT_DEC)
    bl = br.cell(row=BLUME_R, column=1, value="Beta (Blume-adjusted) ★")
    bl.font = Font(bold=True); bl.fill = PatternFill("solid", start_color=_GOLD); bl.border = _BORDER
    bv = br.cell(row=BLUME_R, column=2, value=f"=B{RES_HDR + 10}*B{BETA_R}+B{RES_HDR + 11}*1")
    bv.number_format = "0.0000"
    bv.fill = PatternFill("solid", start_color=_NAVY)
    bv.font = Font(size=11, bold=True, color="FFFFFF")
    bv.alignment = Alignment(horizontal="right"); bv.border = _BORDER

    br.column_dimensions["A"].width = 36
    for col in "BCDE": br.column_dimensions[col].width = 14
    br.freeze_panes = "A4"
    return BETA_R, BLUME_R


# ════════════════════════════════════════════════════════════════════
#  Sheet 6 · Cost_of_Capital  (CAPM Re + Rd + WACC, 한국 KTB·ERP 기준)
# ════════════════════════════════════════════════════════════════════

def _build_coc_kr(wb, model, ticker_dg, BETA_R, BLUME_R):
    """
    한국 자본비용 시트.
    미국과의 차이:
      - Rf: 미국 10Y Treasury → 한국 BOK 국고채 10Y (model.rf)
      - Rm: 미국 _RM_DEFAULT → 한국 모형 (model.e_rm, ERP 반영)
      - Total Debt: bs["totalDebt"] → Σ(DEBT_KEYS) on _fs_wide
      - Cash: bs["cashAndCashEquivalents"] → Σ(CASH_KEYS) on _fs_wide
      - Shares: profile.mktCap/price → model.valuation["shares"]
    """
    cc = wb.create_sheet("Cost_of_Capital")
    cc.merge_cells("A1:F1")
    t = cc["A1"]
    t.value = f"{ticker_dg} — 자본비용 (CAPM Re + Rd + WACC, 한국 KTB·ERP 기준)"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    cc.row_dimensions[1].height = 32

    val = model.valuation
    wide = model._fs_wide

    # WACC chain inputs (한국 데이터)
    interest_ttm = 0.0
    if "interest_expense" in wide.columns:
        interest_ttm = float(pd.to_numeric(wide["interest_expense"], errors="coerce").abs().tail(4).sum())

    # 총부채 = Σ(DEBT_KEYS)
    total_debt_series = pd.Series(0.0, index=wide.index)
    for k in _CONFIG["DEBT_KEYS"]:
        if k in wide.columns:
            total_debt_series = total_debt_series + wide[k].fillna(0)
    total_debt = float(total_debt_series.iloc[-1]) if len(total_debt_series) else 0.0

    # 현금 = Σ(CASH_KEYS)
    cash_series = pd.Series(0.0, index=wide.index)
    for k in _CONFIG["CASH_KEYS"]:
        if k in wide.columns:
            cash_series = cash_series + wide[k].fillna(0)
    cash = float(cash_series.iloc[-1]) if len(cash_series) else 0.0

    # Shares & current price (model.valuation 우선)
    shares_use = float(val.get("shares") or 0.0)
    if isinstance(shares_use, float) and np.isnan(shares_use):
        shares_use = 0.0
    current_price = float(val.get("current_price") or 0.0)
    if isinstance(current_price, float) and np.isnan(current_price):
        current_price = 0.0

    tax = float(model.estimate_tax_rate())
    wacc_db = float(val.get("wacc", np.nan)) if val else np.nan

    # 한국 모형 가정값 (인스턴스에서 추출)
    rf_value = float(getattr(model, "rf", 0.035))
    rm_value = float(getattr(model, "e_rm", 0.085))

    # ── Section A — Re ────────────────────────────────────────
    _section_header(cc, 3, "Section A — Cost of Equity (Re) via CAPM", 6, _TEAL)
    cc.merge_cells("A4:F4")
    cc.cell(row=4, column=1,
            value="  공식:  Re = Rf + Beta × (Rm − Rf)").font = Font(italic=True, size=10, color="595959")

    beta_raw_ref   = f"=Beta_Regression!B{BETA_R}"   if BETA_R   else 1.0
    beta_blume_ref = f"=Beta_Regression!B{BLUME_R}"  if BLUME_R  else 1.0

    spec = [
        (6,  "Risk-free rate (Rf, 한국 KTB 10Y)",   rf_value,            _FMT_PCT, "input"),
        (7,  "Expected market return (Rm, KOSPI)",   rm_value,            _FMT_PCT, "input"),
        (8,  "  Market Risk Premium (MRP) = Rm − Rf","=B7-B6",            _FMT_PCT, "comp"),
        (10, "Beta (raw, from Beta_Regression)",     beta_raw_ref,        "0.0000", "comp"),
        (11, "Beta (Blume-adjusted)",                beta_blume_ref,      "0.0000", "comp"),
        (13, "Re (using raw Beta)",                  "=B6+B10*B8",        _FMT_PCT, "comp"),
        (14, "Re (using Blume Beta) ★",              "=B6+B11*B8",        _FMT_PCT, "highlight"),
    ]
    for r, label, val_, fmt, kind in spec:
        lc = cc.cell(row=r, column=1)
        _set_label(lc, label, bold=(kind == "highlight"))
        if kind == "highlight":
            lc.fill = PatternFill("solid", start_color=_GOLD)
        c = cc.cell(row=r, column=2)
        if kind == "input":
            _set_input(c, val_, fmt)
        elif kind == "highlight":
            _set_computed(c, val_, fmt, highlight=True)
        else:
            _set_computed(c, val_, fmt)

    # ── Section B — Rd ────────────────────────────────────────
    _section_header(cc, 17, "Section B — Cost of Debt (Rd)", 6, _TEAL)
    cc.merge_cells("A18:F18")
    cc.cell(row=18, column=1,
            value="  Rd_pretax = 이자비용 / 총부채  →  Rd_aftertax = Rd_pretax × (1 − T)").font = Font(italic=True, size=10, color="595959")
    rd_spec = [
        (20, "이자비용 (TTM, last 4Q)",            interest_ttm,                _FMT_INT, "input"),
        (21, "총부채 (latest quarter)",            total_debt,                  _FMT_INT, "input"),
        (23, "  Rd pre-tax",                       "=IFERROR(B20/B21,0)",       _FMT_PCT, "comp"),
        (24, "유효세율 (T)",                       tax,                         _FMT_PCT, "input"),
        (26, "Rd after-tax = Rd_pretax × (1 − T) ★","=IFERROR(B23*(1-B24),0)",  _FMT_PCT, "highlight"),
    ]
    for r, label, val_, fmt, kind in rd_spec:
        lc = cc.cell(row=r, column=1)
        _set_label(lc, label, bold=(kind == "highlight"))
        if kind == "highlight":
            lc.fill = PatternFill("solid", start_color=_GOLD)
        c = cc.cell(row=r, column=2)
        if kind == "input":     _set_input(c, val_, fmt)
        elif kind == "highlight": _set_computed(c, val_, fmt, highlight=True)
        else:                   _set_computed(c, val_, fmt)

    # ── Section C — Capital Structure ─────────────────────────
    _section_header(cc, 29, "Section C — 자본구조 (시장가치 기준)", 6, _TEAL)
    # ★ v2.2 옵션 B: Net Debt 는 model.valuation['net_debt'] 직접 입력
    #   (Total Debt / Cash 가 BS 컬럼 매칭 실패로 0 이 되더라도 정합성 유지)
    net_debt_model = float(val.get("net_debt", 0.0)) if val else 0.0
    if isinstance(net_debt_model, float) and np.isnan(net_debt_model):
        net_debt_model = 0.0
    cs_spec = [
        (31, "현재 주가",                           current_price,        _FMT_PRICE, "input"),
        (32, "발행주식수",                           shares_use,           _FMT_INT,   "input"),
        (33, "  시가총액 (E) = 주가 × 주식수",      "=B31*B32",           _FMT_INT,   "comp"),
        (34, "  총부채 (D, B21 링크) — BS 합산",   "=B21",               _FMT_INT,   "comp"),
        (35, "  총자본 (V) = E + D",                "=B33+B34",           _FMT_INT,   "comp"),
        (37, "  E/V (자기자본 비중)",               "=IFERROR(B33/B35,0)",_FMT_PCT,   "comp"),
        (38, "  D/V (부채 비중)",                   "=IFERROR(B34/B35,0)",_FMT_PCT,   "comp"),
        (41, "현금성 자산 — BS 합산",                cash,                 _FMT_INT,   "input"),
        (40, "Net Debt ★ (model.valuation 값으로 고정)", net_debt_model, _FMT_INT, "highlight"),
    ]
    for r, label, val_, fmt, kind in cs_spec:
        lc = cc.cell(row=r, column=1)
        _set_label(lc, label, bold=(kind == "highlight"))
        if kind == "highlight":
            lc.fill = PatternFill("solid", start_color=_GOLD)
        c = cc.cell(row=r, column=2)
        if kind == "input":
            _set_input(c, val_, fmt)
        elif kind == "highlight":
            _set_computed(c, val_, fmt, highlight=True)
        else:
            _set_computed(c, val_, fmt)
    # 안내 메모
    cc.merge_cells("C40:F40")
    note_nd = cc.cell(row=40, column=3,
                      value="  ※ B34-B41 (BS 합산) 가 아닌 model 결과로 고정")
    note_nd.font = Font(italic=True, size=9, color="595959")

    # ── Section D — WACC ──────────────────────────────────────
    # ★ v2.2 옵션 B: WACC 도 model.valuation['wacc'] 로 직접 고정
    #   B47, B48 은 sub-component sanity check 용으로만 표시.
    #   (Beta=1 fallback / Total Debt=0 등으로 Excel 수식 결과가 model 과
    #    어긋나도 최종 WACC 는 model 결과를 우선)
    _section_header(cc, 44, "Section D — WACC", 6, _TEAL)
    cc.merge_cells("A45:F45")
    cc.cell(row=45, column=1,
            value="  WACC = (E/V) × Re + (D/V) × Rd × (1 − T)  "
                  "[참고용: B47+B48; 실제 사용은 B50 = model 결과]"
            ).font = Font(italic=True, size=10, color="595959")
    for r, label, val_, fmt, kind in [
        (47, "  자기자본 기여: (E/V) × Re_Blume (참고용)",   "=B37*B14", _FMT_PCT, "comp"),
        (48, "  부채 기여: (D/V) × Rd_aftertax (참고용)",    "=B38*B26", _FMT_PCT, "comp"),
        (49, "    (Excel sub-component 합 = B47+B48)",       "=B47+B48", _FMT_PCT, "comp"),
    ]:
        lc = cc.cell(row=r, column=1)
        _set_label(lc, label)
        c = cc.cell(row=r, column=2)
        _set_computed(c, val_, fmt)

    # WACC ★ — model 값 직접 입력 (DCF 시트가 참조하는 셀)
    wacc_use = float(val.get("wacc", 0.10)) if val else 0.10
    if isinstance(wacc_use, float) and np.isnan(wacc_use):
        wacc_use = 0.10
    lc = cc.cell(row=50, column=1)
    _set_label(lc, "WACC ★ (model.valuation 값으로 고정)", bold=True)
    lc.fill = PatternFill("solid", start_color=_GOLD)
    cw = cc.cell(row=50, column=2)
    _set_computed(cw, wacc_use, _FMT_PCT, final=True)
    # 안내 메모
    cc.merge_cells("C50:F50")
    note_w = cc.cell(row=50, column=3,
                     value="  ※ B47+B48 (E/V·Re + D/V·Rd) 가 아닌 model 결과로 고정")
    note_w.font = Font(italic=True, size=9, color="595959")

    # ── Section E — Reference (옵션 B 적용으로 Δ = 0 보장) ─────────────
    _section_header(cc, 53, "Section E — 참고: Excel sub-component vs model WACC", 6, _GOLD)
    _set_label(cc.cell(row=55, column=1), "WACC (model.valuation['wacc']) — B50 과 동일")
    if not (isinstance(wacc_db, float) and np.isnan(wacc_db)):
        wc = cc.cell(row=55, column=2, value=float(wacc_db))
        wc.number_format = _FMT_PCT
        wc.fill = PatternFill("solid", start_color=_PHASE1_BG)
        wc.font = Font(bold=True); wc.alignment = Alignment(horizontal="right"); wc.border = _BORDER
    else:
        cc.cell(row=55, column=2, value="N/A").alignment = Alignment(horizontal="right")
    _set_label(cc.cell(row=56, column=1), "  Δ (Excel sub-component sum − model WACC)")
    cc.cell(row=56, column=2, value="=B49-B55").number_format = _FMT_PCT
    cc.cell(row=56, column=2).font = Font(italic=True, color="595959")
    cc.cell(row=56, column=2).border = _BORDER

    cc.merge_cells("A58:F58")
    cc.cell(row=58, column=1,
            value="※ 옵션 B (v2.2): Net Debt(B40), WACC(B50) 는 model 결과로 직접 고정 → "
                  "DCF_Valuation 의 Excel TP = model TP 항상 일치 보장."
            ).font = Font(italic=True, size=9, color="595959")
    cc.merge_cells("A59:F59")
    cc.cell(row=59, column=1,
            value="   Δ(B56) 는 Excel 수식 체인 (E/V·Re + D/V·Rd) 와 model WACC 차이 — "
                  "Beta=1 fallback / BS 매칭 실패 진단용."
            ).font = Font(italic=True, size=9, color="595959")

    cc.column_dimensions["A"].width = 56
    cc.column_dimensions["B"].width = 26
    for col in "CDEF": cc.column_dimensions[col].width = 14


# ════════════════════════════════════════════════════════════════════
#  Sheet 7 · DCF_Valuation  (Y1-17 PV + TV → 적정주가)
# ════════════════════════════════════════════════════════════════════

def _build_dcf_kr(wb, model, ticker_dg):
    dv = wb.create_sheet("DCF_Valuation")
    dv.merge_cells("A1:F1")
    t = dv["A1"]
    t.value = f"{ticker_dg} — 3-Stage DCF: WACC 할인 FCFF → 적정주가 (단위: 원)"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    dv.row_dimensions[1].height = 32

    _section_header(dv, 3, "Inputs (Cost_of_Capital, Phase2_LongTerm 시트 링크)", 6, _TEAL)
    for r, label, formula, fmt in [
        (5, "WACC",                        "=Cost_of_Capital!B50", _FMT_PCT),
        (6, "Terminal growth (g_term)",    "=Phase2_LongTerm!B6",  _FMT_PCT),
        (7, "Net Debt",                    "=Cost_of_Capital!B40", _FMT_INT),
        (8, "발행주식수",                   "=Cost_of_Capital!B32", _FMT_INT),
        (9, "현재 주가",                    "=Cost_of_Capital!B31", _FMT_PRICE),
    ]:
        _set_label(dv.cell(row=r, column=1), label)
        c = dv.cell(row=r, column=2, value=formula)
        _set_computed(c, formula, fmt, highlight=(r == 5))

    _section_header(dv, 12, "연도별 FCFF 할인 (Year 1-17)", 6, _TEAL)
    for c, h in enumerate(["Year","Cal. Year","Phase","FCFF (annual)","Discount Factor","PV of FCFF"], 1):
        cell = dv.cell(row=13, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    try:
        BASE_YEAR = int(str(model.result_df["quarter"].iloc[0])[:4])
    except Exception:
        BASE_YEAR = datetime.now().year

    for yr in range(1, 18):
        r = 13 + yr
        dv.cell(row=r, column=1, value=yr).alignment = Alignment(horizontal="center")
        dv.cell(row=r, column=2, value=BASE_YEAR - 1 + yr).alignment = Alignment(horizontal="center")
        dv.cell(row=r, column=3, value=f"=Phase2_LongTerm!C{r}").alignment = Alignment(horizontal="center")
        c4 = dv.cell(row=r, column=4, value=f"=Phase2_LongTerm!E{r}"); c4.number_format = _FMT_INT
        c5 = dv.cell(row=r, column=5, value=f"=1/(1+$B$5)^A{r}");      c5.number_format = "0.0000"
        c6 = dv.cell(row=r, column=6, value=f"=D{r}*E{r}");              c6.number_format = _FMT_INT
        bg = "E8F4F8" if yr <= 2 else "FFF4D6"
        for cc_idx in range(1, 7):
            cc = dv.cell(row=r, column=cc_idx)
            cc.border = _BORDER
            cc.fill = PatternFill("solid", start_color=bg)

    dv.merge_cells("A32:E32")
    _set_label(dv.cell(row=32, column=1), "  Σ PV of FCFF (Year 1-17)", bold=True)
    sumc = dv.cell(row=32, column=6, value="=SUM(F14:F30)")
    sumc.number_format = _FMT_INT
    sumc.fill = PatternFill("solid", start_color=_LINK_BG)
    sumc.font = Font(size=11, bold=True)
    sumc.alignment = Alignment(horizontal="right"); sumc.border = _BORDER

    _section_header(dv, 35, "Terminal Value (Phase 3, Gordon Growth Model)", 6, _TEAL)
    dv.merge_cells("A36:F36")
    dv.cell(row=36, column=1,
            value="  TV = FCFF_17 × (1 + g) / (WACC − g)  ;  PV = TV / (1+WACC)^17").font = Font(italic=True, size=10, color="595959")

    for r, label, formula, fmt, hl in [
        (38, "FCFF at end of Year 17",                       "=D30",                  _FMT_INT, False),
        (39, "  Terminal Value = FCFF_17 × (1+g)/(WACC−g)",  "=B38*(1+B6)/(B5-B6)",   _FMT_INT, False),
        (40, "  Discount factor for Year 17",                "=1/(1+B5)^17",          "0.0000", False),
        (41, "PV of Terminal Value ★",                       "=B39*B40",              _FMT_INT, True),
    ]:
        lc = dv.cell(row=r, column=1)
        _set_label(lc, label, bold=hl)
        if hl: lc.fill = PatternFill("solid", start_color=_GOLD)
        c = dv.cell(row=r, column=2, value=formula)
        _set_computed(c, formula, fmt, highlight=hl)

    _section_header(dv, 44, "Enterprise Value → Equity Value → 적정주가", 6, _GOLD)
    for r, label, formula, fmt, kind in [
        (46, "Enterprise Value (EV)",                "=F32+B41",                _FMT_INT,  "highlight"),
        (47, "  Equity Value = EV − Net Debt",       "=B46-B7",                 _FMT_INT,  "comp"),
        (49, "적정주가 ★★★",                        "=IFERROR(B47/B8, 0)",     _FMT_PRICE,"final"),
        (50, "  현재 주가",                          "=B9",                     _FMT_PRICE,"comp"),
        (51, "  Upside / (Downside)",                "=IFERROR((B49-B50)/B50,0)",_FMT_PCT, "highlight"),
        (53, "(Implied Market Cap)",                 "=B49*B8",                 _FMT_INT,  "comp"),
    ]:
        lc = dv.cell(row=r, column=1)
        _set_label(lc, label, bold=(kind in ("highlight", "final")))
        if kind == "final":
            lc.fill = PatternFill("solid", start_color=_GOLD)
        c = dv.cell(row=r, column=2, value=formula)
        if kind == "final":
            _set_computed(c, formula, fmt, final=True)
            c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        else:
            _set_computed(c, formula, fmt, highlight=(kind == "highlight"))

    # Reference (model 의 target_price) — 옵션 B 적용으로 Δ = 0 보장
    _section_header(dv, 56, "Reference: 적정주가 (KoreaDCFModel batch 결과)", 6, _GOLD)
    val_dict = model.valuation
    _set_label(dv.cell(row=58, column=1), "TP (model.valuation['target_price'])")
    tp_db = val_dict.get("target_price", np.nan)
    if tp_db is not None and not (isinstance(tp_db, float) and np.isnan(tp_db)):
        tp_c = dv.cell(row=58, column=2, value=float(tp_db))
        tp_c.number_format = _FMT_PRICE
        tp_c.fill = PatternFill("solid", start_color=_PHASE1_BG)
        tp_c.font = Font(bold=True); tp_c.alignment = Alignment(horizontal="right"); tp_c.border = _BORDER
    else:
        dv.cell(row=58, column=2, value="N/A").alignment = Alignment(horizontal="right")
    _set_label(dv.cell(row=59, column=1), "  Δ (Excel TP − model TP)")
    dv.cell(row=59, column=2, value="=B49-B58").number_format = _FMT_PRICE
    dv.cell(row=59, column=2).font = Font(italic=True, color="595959")
    dv.cell(row=59, column=2).border = _BORDER

    # 옵션 B 안내
    dv.merge_cells("A61:F61")
    dv.cell(row=61, column=1,
            value="※ 옵션 B (v2.2): Cost_of_Capital!B40 (Net Debt), B50 (WACC) 가 model 값으로 "
                  "고정되어 있으므로 Δ(B59) 는 항상 0 (반올림 오차 ≤ 1원).").font = Font(italic=True, size=9, color="595959")
    dv.merge_cells("A62:F62")
    dv.cell(row=62, column=1,
            value="   What-if 분석을 원하시면 Phase2_LongTerm 의 ρ/g₀/g_term/horizon (B4:B7) 만 "
                  "조정하세요. WACC 변경은 Cost_of_Capital!B50 직접 수정.").font = Font(italic=True, size=9, color="595959")

    dv.column_dimensions["A"].width = 42
    dv.column_dimensions["B"].width = 26
    for col in "CDEF": dv.column_dimensions[col].width = 18
    dv.freeze_panes = "A14"


# ════════════════════════════════════════════════════════════════════
#  Sheet 8 · Sensitivity_TP  (3개 grid: WACC×g, Beta×g, Re×g)
# ════════════════════════════════════════════════════════════════════

def _build_sens_kr(wb, ticker_dg):
    sn = wb.create_sheet("Sensitivity_TP")
    sn.merge_cells("A1:H1")
    t = sn["A1"]
    t.value = f"{ticker_dg} — 적정주가 민감도 분석 (단위: 원)"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    sn.row_dimensions[1].height = 32

    _section_header(sn, 3, "Reference Values (linked)", 8, _TEAL)
    for r, label, formula, fmt in [
        (5,  "WACC (base)",                 "=Cost_of_Capital!B50", _FMT_PCT),
        (6,  "Re_Blume (base)",             "=Cost_of_Capital!B14", _FMT_PCT),
        (7,  "Beta_Blume (base)",           "=Cost_of_Capital!B11", "0.0000"),
        (8,  "Risk-free rate (Rf)",         "=Cost_of_Capital!B6",  _FMT_PCT),
        (9,  "Expected market return (Rm)", "=Cost_of_Capital!B7",  _FMT_PCT),
        (10, "E/V",                         "=Cost_of_Capital!B37", _FMT_PCT),
        (11, "D/V",                         "=Cost_of_Capital!B38", _FMT_PCT),
        (12, "Rd after-tax",                "=Cost_of_Capital!B26", _FMT_PCT),
        (13, "FCFF Year 17",                "=Phase2_LongTerm!E30", _FMT_INT),
        (14, "Net Debt",                    "=Cost_of_Capital!B40", _FMT_INT),
        (15, "발행주식수",                   "=Cost_of_Capital!B32", _FMT_INT),
    ]:
        _set_label(sn.cell(row=r, column=1), label)
        c = sn.cell(row=r, column=2, value=formula)
        _set_computed(c, formula, fmt)

    g_grid = [0.02, 0.03, 0.04, 0.05, 0.06]
    sumprod_f = 'SUMPRODUCT(Phase2_LongTerm!$E$14:$E$30/(1+{w})^ROW(INDIRECT("1:17")))'

    # ── Grid 1: WACC × g_term ─────────────────────────────────
    _section_header(sn, 18, "Grid 1: TP = f(WACC × g_term)", 8, _GOLD)
    for j, g in enumerate(g_grid):
        c = sn.cell(row=20, column=3 + j, value=g); c.number_format = _FMT_PCT
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", start_color=_NAVY)
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
    for i, off in enumerate([-0.02, -0.01, 0.00, 0.01, 0.02]):
        r = 21 + i
        formula_row = "=$B$5" if off == 0 else f"=$B$5{off:+}"
        c = sn.cell(row=r, column=1, value=formula_row); c.number_format = _FMT_PCT
        c.font = Font(bold=(off == 0), color="FFFFFF")
        c.fill = PatternFill("solid", start_color=_NAVY if off == 0 else "455A64")
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
        for j in range(5):
            col = 3 + j
            wacc_ref = f"$A{r}"; g_ref = f"{get_column_letter(col)}$20"
            sp = sumprod_f.format(w=wacc_ref)
            formula = (f"=({sp}+$B$13*(1+{g_ref})/({wacc_ref}-{g_ref})/(1+{wacc_ref})^17"
                       f"-$B$14)/$B$15")
            cell = sn.cell(row=r, column=col, value=formula)
            cell.number_format = _FMT_PRICE
            cell.alignment = Alignment(horizontal="right"); cell.border = _BORDER
            cell.fill = PatternFill("solid", start_color=_GOLD if (i == 2 and j == 2) else "FFFFFF")

    # ── Grid 2: Beta × g_term (WACC 재계산) ───────────────────
    _section_header(sn, 30, "Grid 2: TP = f(Beta × g_term, WACC 재계산)", 8, _GOLD)
    for j, g in enumerate(g_grid):
        c = sn.cell(row=33, column=3 + j, value=g); c.number_format = _FMT_PCT
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", start_color=_NAVY)
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
    for i, off in enumerate([-0.25, -0.10, 0.00, 0.10, 0.25]):
        r = 34 + i
        formula_row = "=$B$7" if off == 0 else f"=$B$7{off:+}"
        c = sn.cell(row=r, column=1, value=formula_row); c.number_format = "0.0000"
        c.font = Font(bold=(off == 0), color="FFFFFF")
        c.fill = PatternFill("solid", start_color=_NAVY if off == 0 else "455A64")
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
        for j in range(5):
            col = 3 + j
            beta_ref = f"$A{r}"; g_ref = f"{get_column_letter(col)}$33"
            re_local   = f"($B$8+{beta_ref}*($B$9-$B$8))"
            wacc_local = f"($B$10*{re_local}+$B$11*$B$12)"
            sp = sumprod_f.format(w=wacc_local)
            formula = (f"=({sp}+$B$13*(1+{g_ref})/({wacc_local}-{g_ref})/(1+{wacc_local})^17"
                       f"-$B$14)/$B$15")
            cell = sn.cell(row=r, column=col, value=formula)
            cell.number_format = _FMT_PRICE
            cell.alignment = Alignment(horizontal="right"); cell.border = _BORDER
            cell.fill = PatternFill("solid", start_color=_GOLD if (i == 2 and j == 2) else "FFFFFF")

    # ── Grid 3: Re × g_term ───────────────────────────────────
    _section_header(sn, 43, "Grid 3: TP = f(Re × g_term)", 8, _GOLD)
    sn.merge_cells("A44:H44")
    sn.cell(row=44, column=1,
            value="  각 셀이 Re 를 직접 사용 → WACC 재계산. Rd 는 고정.").font = Font(italic=True, size=9, color="595959")
    for j, g in enumerate(g_grid):
        c = sn.cell(row=46, column=3 + j, value=g); c.number_format = _FMT_PCT
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", start_color=_NAVY)
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
    for i, off in enumerate([-0.02, -0.01, 0.00, 0.01, 0.02]):
        r = 47 + i
        formula_row = "=$B$6" if off == 0 else f"=$B$6{off:+}"
        c = sn.cell(row=r, column=1, value=formula_row); c.number_format = _FMT_PCT
        c.font = Font(bold=(off == 0), color="FFFFFF")
        c.fill = PatternFill("solid", start_color=_NAVY if off == 0 else "455A64")
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
        for j in range(5):
            col = 3 + j
            re_ref = f"$A{r}"; g_ref = f"{get_column_letter(col)}$46"
            wacc_local = f"($B$10*{re_ref}+$B$11*$B$12)"
            sp = sumprod_f.format(w=wacc_local)
            formula = (f"=({sp}+$B$13*(1+{g_ref})/({wacc_local}-{g_ref})/(1+{wacc_local})^17"
                       f"-$B$14)/$B$15")
            cell = sn.cell(row=r, column=col, value=formula)
            cell.number_format = _FMT_PRICE
            cell.alignment = Alignment(horizontal="right"); cell.border = _BORDER
            cell.fill = PatternFill("solid", start_color=_GOLD if (i == 2 and j == 2) else "FFFFFF")

    sn.column_dimensions["A"].width = 32
    for col in "BCDEFGH": sn.column_dimensions[col].width = 16


# ════════════════════════════════════════════════════════════════════
#  Sheet 9 · Notes  (방법론 / 가정 / 출처)
# ════════════════════════════════════════════════════════════════════

def _build_notes_kr(wb, model, ticker_dg):
    notes = wb.create_sheet("Notes")
    val = model.valuation or {}
    g0_method = val.get("g0_method", "n/a")
    notes_text = [
        (f"{ticker_dg} — Korea FCFF DCF v2 Excel Export (run: {datetime.now():%Y-%m-%d %H:%M})",
         True, 14, _NAVY, "FFFFFF"),
        ("", False, 11, None, None),
        ("Source: 한국 재무 DB (korea_fs_data_from_DG) + KoreaDCFModel.run() in-memory", False, 11, None, None),
        ("", False, 11, None, None),
        ("정의", True, 12, _TEAL, "FFFFFF"),
        ("  Sales       = revenue (손익계산서)", False, 11, None, None),
        ("  EBIT        = operating_income", False, 11, None, None),
        ("  OPM         = EBIT / Sales", False, 11, None, None),
        ("  NOPAT       = EBIT × (1 − 유효세율)", False, 11, None, None),
        ("  D&A         = da_cf + intangible_amort_cf (현금흐름표)", False, 11, None, None),
        ("  CapEx       = |capex_tangible + capex_intangible| (현금흐름표)", False, 11, None, None),
        ("  NWC         = (유동자산 − 현금) − (유동부채 − 단기차입금)", False, 11, None, None),
        ("  ΔNWC        = NWC(t) − NWC(t−1)  ★ v8 시드: prev_nwc = γ × first/last sales", False, 11, None, None),
        ("  FCFF        = NOPAT + D&A − CapEx − ΔNWC", False, 11, None, None),
        ("", False, 11, None, None),
        ("Phase 구조", True, 12, _TEAL, "FFFFFF"),
        ("  Phase 1 (Year 1-2):  DB forecast 8Q (Ensemble→SARIMA→ETS→Theta) → 분기 합산 연간화", False, 11, None, None),
        ("  Phase 2 (Year 3+):   AR(1) decay g(t) = g_term + (g(t-1) − g_term) × ρ", False, 11, None, None),
        ("                       기간(n_years) 은 EVA spread → Moat 등급에서 결정", False, 11, None, None),
        ("                       ρ 는 EVA 기반 + OLS 추정 가중평균 (각 50%, clip 0.40~0.92)", False, 11, None, None),
        ("  Phase 3 (Year n+):   Gordon Growth — TV = FCFF × (1+g_term)/(WACC − g_term)", False, 11, None, None),
        ("", False, 11, None, None),
        ("g₀ Patch (v9.2 한국 이식) — 음수 처리 + min(forecast, hist)", True, 12, _GOLD, "000000"),
        (f"  현재 종목의 g₀ method: {g0_method}", False, 11, None, None),
        ("  · 과거 시계열 기반 (g0_hist):", False, 11, None, None),
        ("    Method 1: 8Q CAGR  /  Method 2: 12Q TTM (fallback)", False, 11, None, None),
        ("  · forecast 기반 (g0_fc):", False, 11, None, None),
        ("    Method 3: Phase 1 yr1 → yr2 성장률", False, 11, None, None),
        ("  · 결합 규칙:", False, 11, None, None),
        ("    양수 케이스 → min(g0_hist, g0_fc)  (보수적 채택)", False, 11, None, None),
        ("    forecast 음수 → forecast 우선  (사이클 신호 신뢰)", False, 11, None, None),
        ("    부호 불일치 → 더 작은 값 채택", False, 11, None, None),
        ("  · 하한: -20% (사이클릭 침체 허용), 상한: +50%", False, 11, None, None),
        ("", False, 11, None, None),
        ("한국 시장 가정값 (미국과 다른 부분)", True, 12, _GOLD, "000000"),
        ("  · Rf : 한국 BOK 국고채 10Y (Cell 4 시장 파라미터에서 캐싱)", False, 11, None, None),
        ("  · Rm : KOSPI 장기 기대수익률 (ERP 반영)", False, 11, None, None),
        ("  · 시가총액: ks_listed_company_daily_marketcap (백만원 → 원 환산)", False, 11, None, None),
        ("  · 재무 데이터: 천원 → 원 환산 (FS_UNIT_MULTIPLIER)", False, 11, None, None),
        ("  · 단위 환산 후 NWC 시드도 원 단위 정합 유지", False, 11, None, None),
        ("", False, 11, None, None),
        ("Yellow cells (입력) — 변경 시 자동 재계산", True, 11, _GOLD, "000000"),
        ("  Phase2_LongTerm: B4:B7 (ρ, g₀, g_term, 기간) — 유일한 What-if 영역", False, 11, None, None),
        ("  Cost_of_Capital: B6,B7 (Rf/Rm), B20,B24,B31,B32,B41 (interest/tax/price/shares/cash)", False, 11, None, None),
        ("  ※ Main 시트 M11:M15, N4:N7 은 v2.3 부터 참고용 (Forecast 영역에 영향 없음)", False, 11, None, None),
        ("", False, 11, None, None),
        ("★ 옵션 B (v2.2) — model 결과 강제 일치 (Excel TP = model TP 보장)", True, 12, _GOLD, "000000"),
        ("  · Cost_of_Capital!B40 (Net Debt) = model.valuation['net_debt'] 직접 입력 (수식 아님)", False, 11, None, None),
        ("  · Cost_of_Capital!B50 (WACC)     = model.valuation['wacc']     직접 입력 (수식 아님)", False, 11, None, None),
        ("  · 결과: DCF_Valuation Δ(B59) = 0 보장, Sensitivity 의 base 셀도 model 정합", False, 11, None, None),
        ("  · 참고: B47, B48 (E/V·Re + D/V·Rd 합) 은 진단용으로 유지", False, 11, None, None),
        ("", False, 11, None, None),
        ("★ Main 시트 정합성 패치 (v2.3) — Main = Phase2 보장", True, 12, _GOLD, "000000"),
        ("  · Main 시트 forecast (R43-R50) 모든 컬럼 (Sales/OPM/EBIT/.../FCFF) 을", False, 11, None, None),
        ("    model.result_df 에서 직접 가져옴 (수식 아님)", False, 11, None, None),
        ("  · 결과: Main!K46 (2026 합계) = Phase2!E14 (Year 1) 자동 일치", False, 11, None, None),
        ("           Main!K50 (2027 합계) = Phase2!E15 (Year 2) 자동 일치", False, 11, None, None),
        ("  · 이전 (v2.2 까지): Main 의 forecast 가 평균 OPM (M11) 적용한 Excel 수식이라", False, 11, None, None),
        ("    분기별 정밀 계산이 반영되지 않아 Phase2 (model.ph1_annual) 와 어긋났음", False, 11, None, None),
        ("  · What-if 분석 위치: Phase2_LongTerm (ρ/g_term/horizon) + Sensitivity 셀", False, 11, None, None),
        ("", False, 11, None, None),
        ("주의사항", True, 12, _TEAL, "FFFFFF"),
        ("  • 마진 일정 가정 (constant margin) — 실제 DCF 는 단계적 margin 수렴 권고 (Damodaran)", False, 11, None, None),
        ("  • Phase 2 AR(1) 음수 g₀ 시 decay 가 g_term 으로 자연 회복 — 평균 회귀 (Fama-French 1995)", False, 11, None, None),
        ("  • DCF_Valuation 의 Δ(B59) 는 옵션 B 적용으로 항상 0. Cost_of_Capital Δ(B56) 만 진단 의미.", False, 11, None, None),
        ("  • 일부 BS 필드 (cash, 단기차입 등) 가 0 일 수 있음 → 노란 셀 직접 입력 권장", False, 11, None, None),
    ]
    for i, (txt, bold, size, fc, font_c) in enumerate(notes_text, 1):
        c = notes.cell(row=i, column=1, value=txt)
        c.font = Font(name="Calibri", size=size, bold=bold,
                      color=font_c if font_c else "000000")
        if fc: c.fill = PatternFill("solid", start_color=fc)
        c.alignment = Alignment(horizontal="left", vertical="center")
    notes.column_dimensions["A"].width = 130


# ════════════════════════════════════════════════════════════════════
#  Main entry: model.export_korea_excel(out_dir)
# ════════════════════════════════════════════════════════════════════

def export_korea_excel(self, out_dir=None, fetch_extras=True) -> Path:
    """
    KoreaDCFModel.run() 직후 호출. 9-sheet workbook 생성 후 저장.

    Returns
    -------
    Path : 저장된 .xlsx 파일 경로
    """
    if self.valuation is None or self.result_df is None:
        raise RuntimeError(
            f"[{self.ticker_dg}] export_korea_excel 호출 전에 model.run() 을 먼저 실행하세요."
        )

    # 출력 경로
    if out_dir is None:
        out_dir = Path(r"C:/reports") if os.name == "nt" else Path.home() / "reports"
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{self.ticker_dg}_FCFF_Korea_v2.xlsx"

    # 추가 fetch
    if fetch_extras:
        self._fetch_korea_extras(years=10)

    if self.verbose:
        _log(self.ticker_dg, f"Korea Excel 빌드 시작 → {out_file}")

    # 과거 40Q 구성요소 + 예측 8Q
    hist_df = _build_hist_components_kr(self, n_qtrs=40)
    if hist_df.empty:
        raise RuntimeError(f"[{self.ticker_dg}] 과거 FCFF 구성요소 생성 실패")
    fc_df = self.result_df.copy()
    if "quarter" not in fc_df.columns:
        fc_df["quarter"] = fc_df["date"].apply(lambda d: f"{d.year}Q{d.quarter}")

    wb = Workbook()
    _build_main_kr(self, wb, hist_df, fc_df, self.ticker_dg)            # 1
    _build_verification_kr(wb, hist_df, self.ticker_dg)                 # 2
    _build_phase2_kr(wb, self, self.ticker_dg)                          # 3
    _build_moat_kr(wb, self, self.ticker_dg)                            # 4
    BETA_R, BLUME_R = _build_beta_kr(wb, self, self.ticker_dg)          # 5
    _build_coc_kr(wb, self, self.ticker_dg, BETA_R, BLUME_R)            # 6
    _build_dcf_kr(wb, self, self.ticker_dg)                             # 7
    _build_sens_kr(wb, self.ticker_dg)                                  # 8
    _build_notes_kr(wb, self, self.ticker_dg)                           # 9

    wb.save(out_file)
    if self.verbose:
        _log(self.ticker_dg, f"✓ Korea Excel 저장 완료 ({len(wb.sheetnames)} sheets)")

    return out_file



# ════════════════════════════════════════════════════════════════════
#  공개 등록 함수 — KoreaDCFModel 클래스에 메서드 attach
# ════════════════════════════════════════════════════════════════════

def register_excel_export(cls, verbose: bool = True):
    """
    KoreaDCFModel 클래스에 _fetch_korea_extras / export_korea_excel attach.

    Parameters
    ----------
    cls : type
        KoreaDCFModel 클래스 (또는 호환 클래스).
    verbose : bool, default True
        등록 메시지 출력 여부.

    Notes
    -----
    호출 후 cls 인스턴스에서 다음 메서드 사용 가능:
      - model._fetch_korea_extras(years=10)
      - model.export_korea_excel(out_dir=None, fetch_extras=True)
    """
    cls._fetch_korea_extras = _fetch_korea_extras
    cls.export_korea_excel  = export_korea_excel

    if verbose:
        print(f"[OK] {cls.__name__} patched via FCFF_Result_export_excel_module:")
        print("     ① _fetch_korea_extras   (10Y monthly close: 종목 + KOSPI)")
        print("     ② export_korea_excel    (9-sheet workbook)")


__all__ = [
    "configure",
    "register_excel_export",
    # 디버깅·고급 사용자용 노출
    "_fetch_korea_extras",
    "export_korea_excel",
    "_build_hist_components_kr",
    "_build_main_kr",
    "_build_verification_kr",
    "_build_phase2_kr",
    "_build_moat_kr",
    "_build_beta_kr",
    "_build_coc_kr",
    "_build_dcf_kr",
    "_build_sens_kr",
    "_build_notes_kr",
]
