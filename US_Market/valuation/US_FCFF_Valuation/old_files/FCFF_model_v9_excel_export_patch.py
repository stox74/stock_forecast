# ═══════════════════════════════════════════════════════════════════════
#  Cell 8.5 (NEW)  ·  DCFModel V10 Excel Export Patch
# ═══════════════════════════════════════════════════════════════════════
#  V9 DCF 노트북의 Cell 8 (DCFModel 클래스 정의) 바로 다음에 새 셀로 추가
#  하세요. 기존 Cell 8 을 수정하지 않고 monkey-patch 방식으로 3가지 기능을
#  DCFModel 에 주입합니다.
#
#  주입 내역
#    ① _estimate_phase2_growth  — 음수 g₀ 처리 패치 (v9 → v9.1)
#    ② _fetch_v10_extras        — 10Y monthly close + profile fetch
#    ③ export_v10_excel         — 9-sheet workbook 빌드 & 저장
#
#  사용 예
#    model = DCFModel(ticker="NVDA", engine=engine, verbose=True).run()
#    out_path = model.export_v10_excel(out_dir="C:/reports")
# ═══════════════════════════════════════════════════════════════════════

import os, time
import requests
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path


# ════════════════════════════════════════════════════════════════════
#  Patch ①  ·  _estimate_phase2_growth  (음수 g₀ 처리, v9.1)
# ════════════════════════════════════════════════════════════════════
#  변경점
#    1) g₀ 하한을 G_TERM(4%) → G_FLOOR_MIN(-20%) 으로 완화
#    2) Method 2 (12Q rolling TTM vs TTM-2) 추가 — 부호 반전 명시 처리
#    3) Method 3 (Phase 1 yr1/yr2) 도 부호 반전 케이스 분리
#    4) 최후 fallback 을 임의의 0.10 → G_TERM 으로 변경
#    5) AR(1) clip 하한을 G_TERM → G_FLOOR_MIN 으로 완화
#       → 음수 g₀ 가 g_term 으로 자연스럽게 회복되는 곡선 표현
#    6) ρ OLS 보정 시 양수 조건 제거 — 음수 annual 도 회귀에 포함
# ════════════════════════════════════════════════════════════════════

def _estimate_phase2_growth_v91(self, wacc: float) -> tuple:
    """
    Phase 2 AR(1) 성장 경로 — 음수 g₀ 허용 + 다중 fallback 체인 (v9.1)
    ─────────────────────────────────────────────────────────────
    반환: (g_path, moat_label, rho, n_years)
    """
    G_TERM       = self.gdp_growth
    G_FLOOR_MIN  = -0.20    # ★ 새 하한 (사이클릭 침체 -20% 까지 허용)
    G_CEIL       = 0.80

    h  = self._fcff_history
    g0 = None
    g0_method = None

    # ── Method 1: 8Q CAGR (기존, 단 G_FLOOR_MIN 으로 완화) ─────────
    if h is not None and len(h.dropna()) >= 8:
        h_c = h.dropna()
        f0, fl = float(h_c.iloc[-8]), float(h_c.iloc[-1])
        if f0 > 0 and fl > 0:
            cagr_q = (fl / f0) ** (4.0 / 8) - 1
            g0 = float(np.clip((1 + cagr_q) ** 4 - 1, G_FLOOR_MIN, G_CEIL))
            g0_method = "8Q-CAGR"

    # ── Method 2: 12Q rolling TTM vs TTM-2 (★ NEW) ──────────────
    if g0 is None and h is not None and len(h.dropna()) >= 12:
        h_c = h.dropna()
        ttm_recent = float(h_c.iloc[-4:].sum())
        ttm_prior  = float(h_c.iloc[-8:-4].sum())
        if abs(ttm_prior) > 1e-6:
            if ttm_prior > 0 and ttm_recent > 0:
                g0 = ttm_recent / ttm_prior - 1
            elif ttm_prior > 0 and ttm_recent <= 0:
                g0 = ttm_recent / ttm_prior - 1            # 자연스럽게 음수
            elif ttm_prior < 0 and ttm_recent < 0:
                g0 = (ttm_recent - ttm_prior) / abs(ttm_prior)  # 적자 변화율
            else:                                          # 적자→흑자
                g0 = G_TERM                                # 보수적
            g0 = float(np.clip(g0, G_FLOOR_MIN, G_CEIL))
            g0_method = "12Q-TTM"

    # ── Method 3: Phase 1 (forecast 8Q) yr1 vs yr2 ────────────
    if g0 is None:
        fcff_q = self.result_df["fcff"].values
        yr1 = float(np.sum(fcff_q[:4]))
        yr2 = float(np.sum(fcff_q[4:8])) if len(fcff_q) >= 8 else yr1
        if abs(yr1) > 1e-6:
            if yr1 > 0 and yr2 > 0:
                g0 = yr2 / yr1 - 1
            elif yr1 > 0 and yr2 <= 0:
                g0 = yr2 / yr1 - 1
            elif yr1 < 0 and yr2 < 0:
                g0 = (yr2 - yr1) / abs(yr1)
            else:                                          # 적자→흑자
                g0 = G_TERM
            g0 = float(np.clip(g0, G_FLOOR_MIN, G_CEIL))
            g0_method = "FC-yr1-yr2"

    # ── Method 4: 최후 fallback ───────────────────────────────
    if g0 is None:
        g0 = G_TERM                # ★ 기존 0.10 → 중립적 G_TERM
        g0_method = "default-G_TERM"

    # ── EVA → 해자 등급 → (ρ, 기간) ────────────────────────────
    eva  = self._compute_eva_spread(wacc)
    rho, n_years, moat_label = self._moat_to_rho_and_years(eva)

    # ── ρ OLS 보정 (★ 양수 조건 제거 — 음수 annual 도 포함) ──
    if h is not None and len(h.dropna()) >= 12:
        h_c = h.dropna()
        annual_vals = [float(h_c.iloc[y:y+4].sum())
                       for y in range(0, len(h_c) - len(h_c) % 4, 4)]
        if len(annual_vals) >= 4:
            g_hist = [(annual_vals[i] / annual_vals[i-1] - 1)
                      if abs(annual_vals[i-1]) > 1e-6 else 0.0
                      for i in range(1, len(annual_vals))]
            if len(g_hist) >= 3:
                yt = np.array([g - G_TERM for g in g_hist[1:]])
                xt = np.array([g - G_TERM for g in g_hist[:-1]])
                if np.dot(xt, xt) > 1e-10:
                    rho_ols = float(np.clip(
                        np.dot(xt, yt) / np.dot(xt, xt),
                        0.40, 0.92
                    ))
                    rho = float(np.clip(0.5 * rho + 0.5 * rho_ols, 0.40, 0.92))
                    if self.verbose:
                        log(self.ticker,
                            f"ρ 보정: EVA기반={rho:.3f}  OLS={rho_ols:.3f}  "
                            f"→ 혼합={rho:.3f}")

    # ── AR(1) 성장 경로 — clip 하한 완화 ──────────────────────
    g = g0
    g_path = []
    for _ in range(n_years):
        g = G_TERM + (g - G_TERM) * rho
        g_path.append(float(np.clip(g, G_FLOOR_MIN, G_CEIL)))

    if self.verbose:
        log(self.ticker,
            f"Phase2 [{moat_label}] ρ={rho:.3f} {n_years}년  "
            f"g0={g0:.1%} ({g0_method}) → {g_path[0]:.1%} ... {g_path[-1]:.1%}")

    self._eva_cache = {
        "eva_spread":  eva.get("eva_spread", np.nan),
        "roic":        eva.get("roic",        np.nan),
        "n_positive":  eva.get("n_positive",  0),
        "moat_label":  moat_label,
        "rho":         rho,
        "n_phase2":    n_years,
        "eva_series":  eva.get("eva_series",  []),
        "g0":          g0,           # ★ 추가: 디버깅용
        "g0_method":   g0_method,    # ★ 추가
    }
    return g_path, moat_label, rho, n_years


DCFModel._estimate_phase2_growth = _estimate_phase2_growth_v91


# ════════════════════════════════════════════════════════════════════
#  Patch ②  ·  _fetch_v10_extras  (Beta_Regression / Cost_of_Capital 용)
# ════════════════════════════════════════════════════════════════════

def _fetch_v10_extras(self, years: int = 10) -> None:
    """
    Excel export 에 필요한 추가 FMP 데이터 fetch.
    설정 후 self._monthly_prices, self._monthly_returns, self._profile 에 저장.
    """
    # 10Y monthly close — TICKER + ^GSPC
    end_dt   = datetime.today()
    start_dt = end_dt - timedelta(days=years * 365 + 30)

    def _hist_monthly(symbol: str) -> pd.Series:
        url = f"{FMP_BASE}/historical-price-full/{symbol}"
        params = {"apikey": self.api_key,
                  "from": start_dt.strftime("%Y-%m-%d"),
                  "to":   end_dt.strftime("%Y-%m-%d")}
        for attempt in range(3):
            try:
                r = requests.get(url, params=params, timeout=60)
                if r.status_code == 429:
                    time.sleep(2 + attempt); continue
                r.raise_for_status()
                hist = r.json().get("historical", [])
                if not hist:
                    return pd.Series(dtype=float)
                df = pd.DataFrame(hist)
                df["date"] = pd.to_datetime(df["date"])
                return (df.sort_values("date")
                          .set_index("date")["close"]
                          .resample("ME").last().dropna()
                          .tail(years * 12))
            except Exception as e:
                if attempt == 2:
                    if self.verbose:
                        log(self.ticker, f"monthly price fetch fail [{symbol}]: {e}")
                    return pd.Series(dtype=float)
                time.sleep(0.5 + attempt * 0.5)
        return pd.Series(dtype=float)

    ticker_monthly = _hist_monthly(self.ticker)
    sp500_monthly  = _hist_monthly("%5EGSPC")
    self._monthly_prices  = pd.DataFrame(
        {self.ticker: ticker_monthly, "SP500": sp500_monthly}
    ).dropna()
    self._monthly_returns = self._monthly_prices.pct_change().dropna()

    # Profile (현재 주가 + shares 도출용)
    self._profile = {}
    try:
        r = requests.get(f"{FMP_BASE}/profile/{self.ticker}",
                         params={"apikey": self.api_key}, timeout=20)
        r.raise_for_status()
        data = r.json()
        if isinstance(data, list) and data:
            self._profile = data[0]
    except Exception as e:
        if self.verbose:
            log(self.ticker, f"profile fetch fail: {e}")

    if self.verbose:
        log(self.ticker,
            f"V10 extras: {len(self._monthly_returns)}개월 returns, "
            f"profile={'OK' if self._profile else 'N/A'}")


DCFModel._fetch_v10_extras = _fetch_v10_extras


# ════════════════════════════════════════════════════════════════════
#  Patch ③  ·  export_v10_excel  (9-sheet workbook 빌드)
# ════════════════════════════════════════════════════════════════════
#  V10 Extractor 의 sheet 구성을 그대로 따른다:
#    1. {TICKER}_FCFF       40Q hist + 8Q forecast (수식 기반)
#    2. Verification        YoY 성장률 교차검증
#    3. Phase2_LongTerm     Year 1-22 AR(1) decay
#    4. Moat_Comparison     4 등급 ρ 비교
#    5. Beta_Regression     10Y monthly returns + SLOPE/INTERCEPT/RSQ
#    6. Cost_of_Capital     CAPM Re + Rd + WACC
#    7. DCF_Valuation       Y1-17 PV + TV → Target Price
#    8. Sensitivity_TP      WACC×g, Beta×g, Re×g
#    9. Notes               방법론 / 가정 / 출처
# ════════════════════════════════════════════════════════════════════

# ── 색상 팔레트 (V10 LD1 과 동일) ────────────────────────────────
_NAVY, _TEAL, _GOLD       = "0D1B2A", "0E7C7B", "F4A261"
_ZEBRA, _FCST_BG, _ASSUMP = "F5F7FA", "FFF4D6", "FFF9C4"
_GREY, _BLUE              = "D9D9D9", "0000FF"
_PHASE1_BG, _PHASE2_BG, _PHASE3_BG = "E8F4F8", "FFF4D6", "F0E5C8"
_LINK_BG = "E0F0FF"

_FMT_INT, _FMT_PCT, _FMT_PCT3, _FMT_DEC, _FMT_PRICE = (
    "#,##0;[Red](#,##0);-", "0.00%", "0.000%", "0.0000", "$#,##0.00"
)
_THIN   = Side(style="thin", color=_GREY)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# ── CAPM 가정값 (DCFModel 에는 없으므로 default) ─────────────────
_RF_DEFAULT, _RM_DEFAULT  = 0.045, 0.100
_BLUME_W_RAW, _BLUME_W_MARKET = 0.67, 0.33


# ── Style helpers ───────────────────────────────────────────────
def _set_label(cell, text, bold=False, italic=False, size=10, color="000000"):
    cell.value = text
    cell.font  = Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)
    cell.alignment = Alignment(horizontal="left")
    cell.border = _BORDER

def _set_input(cell, value, fmt):
    cell.value = value
    cell.number_format = fmt
    cell.fill = PatternFill("solid", start_color=_ASSUMP)
    cell.font = Font(name="Calibri", size=10, color=_BLUE, bold=True)
    cell.alignment = Alignment(horizontal="right")
    cell.border = _BORDER

def _set_computed(cell, value, fmt, highlight=False, final=False):
    cell.value = value
    cell.number_format = fmt
    if final:
        cell.fill = PatternFill("solid", start_color=_NAVY)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    elif highlight:
        cell.fill = PatternFill("solid", start_color=_GOLD)
        cell.font = Font(name="Calibri", size=11, bold=True)
    else:
        cell.fill = PatternFill("solid", start_color=_LINK_BG)
        cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="right")
    cell.border = _BORDER

def _section_header(ws, row, text, span=6, color=_TEAL):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    h = ws.cell(row=row, column=1, value=text)
    h.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", start_color=color)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


# ── 과거 FCFF 구성요소 (Sales/OPM/EBIT/.../FCFF) — 40Q rebuild ────
def _build_hist_components(model, n_qtrs: int = 40) -> pd.DataFrame:
    """
    DCFModel._compute_historical_fcff 와 동일 로직이지만 모든 구성요소
    (Sales, OPM, EBIT, NOPAT, DA, CapEx, NWC, ΔNWC, FCFF) 를 캡처해서
    DataFrame 으로 반환. V10 main sheet 채우기용.
    """
    sales_act = model._sales_actual
    if sales_act is None or sales_act.empty:
        return pd.DataFrame()

    # tax/alpha/beta/gamma 재계산 (idempotent)
    tax = model.estimate_tax_rate()
    alpha, _ = model.estimate_ratio_coef("depreciationAndAmortization", "D&A")
    beta,  _ = model.estimate_ratio_coef("capitalExpenditure", "CapEx", take_abs=True)
    gamma, _ = model.estimate_nwc_coef()

    inc = model._inc.set_index("date")
    rows = []
    first_sales = float(sales_act.iloc[0])
    prev_nwc = gamma * first_sales

    for dt, sales in sales_act.items():
        if pd.isna(sales) or sales <= 0:
            prev_nwc = gamma * sales if pd.notna(sales) else prev_nwc
            continue
        op_v  = inc["operatingIncome"].get(dt, np.nan) if "operatingIncome" in inc.columns else np.nan
        rev_v = inc["revenue"].get(dt, np.nan) if "revenue" in inc.columns else np.nan
        opm = (float(op_v) / float(rev_v)
               if (pd.notna(op_v) and pd.notna(rev_v) and float(rev_v) > 0) else np.nan)
        if pd.isna(opm):
            prev_nwc = gamma * sales; continue
        ebit  = sales * opm
        nopat = ebit * (1 - tax)
        da    = alpha * sales
        capex = beta * sales
        nwc   = gamma * sales
        dnwc  = nwc - prev_nwc
        prev_nwc = nwc
        fcff = nopat + da - capex - dnwc
        rows.append({
            "date": dt,
            "quarter": f"{dt.year}Q{dt.quarter}",
            "Sales": sales, "OPM": opm, "EBIT": ebit, "NOPAT": nopat,
            "DA": da, "CapEx": capex, "NWC": nwc, "ΔNWC": dnwc, "FCFF": fcff,
        })
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows).tail(n_qtrs).reset_index(drop=True)
    return out


# ── Sheet builders (model 인스턴스를 인자로 받음) ─────────────────

def _build_main(model, wb, hist_df, fc_df, ticker):
    ws = wb.active
    ws.title = f"{ticker}_FCFF"

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"{ticker} — Quarterly FCFF Components & 8-Quarter Forecast"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Quarter","Sales","OPM","EBIT","NOPAT","D&A","CapEx","NWC","ΔNWC","FCFF"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_TEAL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
    ws.row_dimensions[2].height = 24

    N_HIST = len(hist_df)

    # 과거 (rows 3..2+N_HIST)
    for i in range(N_HIST):
        rd = hist_df.iloc[i]
        r = i + 3
        ws.cell(row=r, column=1, value=rd["quarter"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=float(rd["Sales"])).number_format = _FMT_INT
        ws.cell(row=r, column=3, value=float(rd["OPM"])).number_format  = _FMT_PCT
        ws.cell(row=r, column=4, value=float(rd["EBIT"])).number_format = _FMT_INT
        ws.cell(row=r, column=5, value=float(rd["NOPAT"])).number_format= _FMT_INT
        ws.cell(row=r, column=6, value=float(rd["DA"])).number_format   = _FMT_INT
        ws.cell(row=r, column=7, value=float(rd["CapEx"])).number_format= _FMT_INT
        ws.cell(row=r, column=8, value=float(rd["NWC"])).number_format  = _FMT_INT
        if i == 0:
            for c_idx in (9, 10):
                cc = ws.cell(row=r, column=c_idx, value="–")
                cc.alignment = Alignment(horizontal="center")
        else:
            ws.cell(row=r, column=9,  value=f"=H{r}-H{r-1}").number_format = _FMT_INT
            ws.cell(row=r, column=10, value=f"=E{r}+F{r}-G{r}-I{r}").number_format = _FMT_INT
        fill = PatternFill("solid", start_color=_ZEBRA) if i % 2 == 0 else None
        for col in range(1, 11):
            cc = ws.cell(row=r, column=col)
            cc.border = _BORDER
            cc.font = Font(name="Calibri", size=10)
            if fill: cc.fill = fill

    # 향후 8Q (수식 — 가정셀 참조)
    fcst_top  = Side(style="medium", color=_GOLD)
    fcst_fill = PatternFill("solid", start_color=_FCST_BG)
    fcst_start_row = 3 + N_HIST

    for j in range(min(8, len(fc_df))):
        fc_label = fc_df.iloc[j].get("quarter", f"F{j+1}")
        try:
            fq_num = int(str(fc_label)[-1])
        except Exception:
            fq_num = ((j % 4) + 1)
        r = fcst_start_row + j
        prev_yr_row = r - 4
        growth_row  = 3 + fq_num   # Q1→4, Q2→5, Q3→6, Q4→7

        ws.cell(row=r, column=1, value=str(fc_label)).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=f"=B{prev_yr_row}*(1+$N${growth_row})").number_format = _FMT_INT
        ws.cell(row=r, column=3, value="=$M$11").number_format = _FMT_PCT
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}").number_format = _FMT_INT
        ws.cell(row=r, column=5, value=f"=D{r}*(1-$M$12)").number_format = _FMT_INT
        ws.cell(row=r, column=6, value=f"=B{r}*$M$13").number_format = _FMT_INT
        ws.cell(row=r, column=7, value=f"=B{r}*$M$14").number_format = _FMT_INT
        ws.cell(row=r, column=8, value=f"=B{r}*$M$15").number_format = _FMT_INT
        ws.cell(row=r, column=9, value=f"=H{r}-H{r-1}").number_format = _FMT_INT
        ws.cell(row=r, column=10, value=f"=E{r}+F{r}-G{r}-I{r}").number_format = _FMT_INT

        for col in range(1, 11):
            cc = ws.cell(row=r, column=col)
            cc.border = Border(left=_THIN, right=_THIN, bottom=_THIN,
                               top=fcst_top if j == 0 else _THIN)
            cc.fill = fcst_fill
            cc.font = Font(name="Calibri", size=10)

    # ── 가정셀 (M11:M15) + Q별 성장률 (N4:N7) ──────────────────
    # YoY geometric mean by quarter number
    def _quarterly_yoy(df):
        g = {1: np.nan, 2: np.nan, 3: np.nan, 4: np.nan}
        df = df.copy()
        df["q_num"] = df["quarter"].astype(str).str[-1].astype(int)
        for q in [1, 2, 3, 4]:
            sq = df.loc[df["q_num"] == q].sort_values("quarter")["Sales"].values
            if len(sq) >= 2:
                yoy = sq[1:] / sq[:-1] - 1
                g[q] = float(np.prod(1 + yoy)) ** (1.0 / len(yoy)) - 1
        return g

    growth = _quarterly_yoy(hist_df)

    # M열 가정값 (오른쪽에 노란 input 셀)
    _set_label(ws.cell(row=10, column=12), "Forecast Assumptions (editable)", bold=True)
    ws.cell(row=10, column=12).fill = PatternFill("solid", start_color=_TEAL)
    ws.cell(row=10, column=12).font = Font(bold=True, color="FFFFFF")

    fc_assumptions = {
        "OPM":       float(fc_df["opm_forecast"].mean()) if "opm_forecast" in fc_df.columns else 0.10,
        "Tax rate":  float(fc_df["tax_rate"].iloc[0])    if "tax_rate"     in fc_df.columns else 0.21,
        "D&A / Sales":   float(fc_df["da"].mean()    / fc_df["sales_forecast"].mean()) if "sales_forecast" in fc_df.columns else 0.05,
        "CapEx / Sales": float(fc_df["capex"].mean()  / fc_df["sales_forecast"].mean()) if "sales_forecast" in fc_df.columns else 0.07,
        "NWC / Sales":   float(fc_df["nwc"].mean()    / fc_df["sales_forecast"].mean()) if "sales_forecast" in fc_df.columns else 0.15,
    }
    for i, (k, v) in enumerate(fc_assumptions.items(), start=11):
        _set_label(ws.cell(row=i, column=12), k)
        _set_input(ws.cell(row=i, column=13), v, _FMT_PCT)

    # N열 분기별 성장률
    _set_label(ws.cell(row=2, column=14), "Quarterly YoY Growth (geo)", bold=True)
    ws.cell(row=2, column=14).fill = PatternFill("solid", start_color=_TEAL)
    ws.cell(row=2, column=14).font = Font(bold=True, color="FFFFFF")
    ws.merge_cells("N2:N3")
    for q in [1, 2, 3, 4]:
        _set_label(ws.cell(row=3 + q, column=13), f"Q{q}")
        _set_input(ws.cell(row=3 + q, column=14),
                   growth[q] if not np.isnan(growth[q]) else 0.0, _FMT_PCT3)

    # 컬럼 폭
    ws.column_dimensions["A"].width = 12
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["L"].width = 22
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 14
    ws.freeze_panes = "B3"


def _build_verification(wb, hist_df, ticker):
    ws = wb.create_sheet("Verification")
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"{ticker} — YoY Growth Cross-Check (geo mean by quarter)"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    headers = ["Quarter Number", "# Observations", "Geo. Mean YoY", "Arith. Mean YoY"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_TEAL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    df = hist_df.copy()
    df["q_num"] = df["quarter"].astype(str).str[-1].astype(int)
    for q in [1, 2, 3, 4]:
        sq = df.loc[df["q_num"] == q].sort_values("quarter")["Sales"].values
        r = 3 + q
        if len(sq) >= 2:
            yoy = sq[1:] / sq[:-1] - 1
            arith = float(np.mean(yoy))
            geo   = float(np.prod(1 + yoy)) ** (1.0 / len(yoy)) - 1
            n = len(yoy)
        else:
            arith, geo, n = np.nan, np.nan, 0
        _set_label(ws.cell(row=r, column=1), f"Q{q}")
        ws.cell(row=r, column=2, value=n).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).border = _BORDER
        if n > 0:
            _set_computed(ws.cell(row=r, column=3), geo,   _FMT_PCT3)
            _set_computed(ws.cell(row=r, column=4), arith, _FMT_PCT3)
        else:
            ws.cell(row=r, column=3, value="N/A").alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=4, value="N/A").alignment = Alignment(horizontal="center")

    for col in "ABCD":
        ws.column_dimensions[col].width = 22


def _build_phase2(wb, model, ticker):
    """Phase2_LongTerm — Year 1-22 (Phase 1 + 2 + 3 padding)."""
    ws = wb.create_sheet("Phase2_LongTerm")

    val = model.valuation
    ph1 = val.get("ph1_annual", [])
    ph2 = val.get("ph2_annual", [])
    g_path = val.get("ph2_growth", [])
    rho     = val.get("moat_rho", 0.83)
    g_term  = val.get("g_terminal", 0.04)
    g0      = getattr(model, "_eva_cache", {}).get("g0", np.nan)
    g0_method = getattr(model, "_eva_cache", {}).get("g0_method", "n/a")
    moat = val.get("moat_label", "N/A")
    n_ph2 = len(ph2)

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"{ticker} — Phase 2 Long-Term FCFF (AR(1) decay → g_term)"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Inputs
    _section_header(ws, 3, "Inputs (editable)", 6, _TEAL)
    _set_label(ws.cell(row=4, column=1), "ρ (AR(1) persistence)")
    _set_input(ws.cell(row=4, column=2), float(rho), _FMT_DEC)
    _set_label(ws.cell(row=5, column=1), "g₀ (initial growth)")
    _set_input(ws.cell(row=5, column=2),
               float(g0) if not np.isnan(g0) else 0.0, _FMT_PCT)
    _set_label(ws.cell(row=6, column=1), "g_term (terminal growth)")
    _set_input(ws.cell(row=6, column=2), float(g_term), _FMT_PCT)
    _set_label(ws.cell(row=7, column=1), "Phase 2 horizon (years)")
    _set_input(ws.cell(row=7, column=2), int(n_ph2), "0")

    _set_label(ws.cell(row=9, column=1), f"Moat:  {moat}    g₀ method:  {g0_method}",
               italic=True, color="595959")

    # Year-by-Year table
    _section_header(ws, 12, "Year-by-Year FCFF Schedule (Year 1-22)", 6, _TEAL)
    for c, h in enumerate(["Year","Cal. Year","Phase","g(t)","FCFF","Notes"], 1):
        cell = ws.cell(row=13, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    # 시작 연도 — 예측 첫 분기에서 추출
    try:
        BASE_YEAR = int(str(model.result_df["quarter"].iloc[0])[:4])
    except Exception:
        BASE_YEAR = datetime.now().year

    ph1_growth = []
    if len(ph1) >= 2 and ph1[0] != 0:
        ph1_growth = [np.nan, ph1[1]/ph1[0] - 1]
    elif len(ph1) >= 1:
        ph1_growth = [np.nan]

    # Year 1-2: Phase 1 (DB forecast)
    for y in range(min(2, len(ph1))):
        r = 14 + y
        ws.cell(row=r, column=1, value=y+1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=BASE_YEAR + y).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value="Phase 1").alignment = Alignment(horizontal="center")
        if y < len(ph1_growth) and not np.isnan(ph1_growth[y]):
            _set_computed(ws.cell(row=r, column=4), float(ph1_growth[y]), _FMT_PCT)
        else:
            ws.cell(row=r, column=4, value="—").alignment = Alignment(horizontal="center")
        _set_computed(ws.cell(row=r, column=5), float(ph1[y]), _FMT_INT)
        ws.cell(row=r, column=6, value="Forecast 8Q sum").font = Font(italic=True, size=9, color="595959")
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=_PHASE1_BG)
            ws.cell(row=r, column=col).border = _BORDER

    # Year 3 ~ Year (2+n_ph2): Phase 2 (AR(1))
    last_fcff = float(ph1[-1]) if ph1 else 1.0
    for j in range(n_ph2):
        y_idx = 2 + j
        r = 14 + y_idx
        ws.cell(row=r, column=1, value=y_idx+1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=BASE_YEAR + y_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value="Phase 2").alignment = Alignment(horizontal="center")
        _set_computed(ws.cell(row=r, column=4), float(g_path[j]) if j < len(g_path) else np.nan, _FMT_PCT)
        _set_computed(ws.cell(row=r, column=5), float(ph2[j]),     _FMT_INT)
        ws.cell(row=r, column=6, value=f"AR(1) decay  ρ={rho:.3f}").font = Font(italic=True, size=9, color="595959")
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=_PHASE2_BG)
            ws.cell(row=r, column=col).border = _BORDER
        last_fcff = float(ph2[j])

    # Phase 3 padding (다음해 ~ Year 22) — TV growth 적용
    last_year_done = 2 + n_ph2
    for y_idx in range(last_year_done, 22):
        r = 14 + y_idx
        last_fcff = last_fcff * (1 + g_term)
        ws.cell(row=r, column=1, value=y_idx+1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=BASE_YEAR + y_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value="Phase 3").alignment = Alignment(horizontal="center")
        _set_computed(ws.cell(row=r, column=4), float(g_term), _FMT_PCT)
        _set_computed(ws.cell(row=r, column=5), last_fcff,     _FMT_INT)
        ws.cell(row=r, column=6, value="Terminal growth").font = Font(italic=True, size=9, color="595959")
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=_PHASE3_BG)
            ws.cell(row=r, column=col).border = _BORDER

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28


def _build_moat(wb, model, ticker):
    ws = wb.create_sheet("Moat_Comparison")
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"{ticker} — Moat Classification & ρ Comparison"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    _section_header(ws, 3, "EVA Spread (current snapshot)", 8, _TEAL)
    val = model.valuation
    eva_cache = getattr(model, "_eva_cache", {})
    rows_eva = [
        ("ROIC (TTM)",                val.get("roic", np.nan),       _FMT_PCT),
        ("WACC",                      val.get("wacc", np.nan),       _FMT_PCT),
        ("EVA Spread = ROIC − WACC",  val.get("eva_spread", np.nan), _FMT_PCT),
        ("Positive quarters / 20Q",   eva_cache.get("n_positive", 0), "0"),
        ("Current Moat label",         val.get("moat_label", "N/A"),  None),
        ("ρ (AR(1) persistence) — selected", val.get("moat_rho", np.nan), _FMT_DEC),
        ("Phase 2 horizon (years)",   val.get("n_phase2", 0),        "0"),
    ]
    for i, (label, v, fmt) in enumerate(rows_eva):
        r = 5 + i
        _set_label(ws.cell(row=r, column=1), label)
        if isinstance(v, str):
            ws.cell(row=r, column=2, value=v).alignment = Alignment(horizontal="right")
        elif fmt is None:
            ws.cell(row=r, column=2, value=v).alignment = Alignment(horizontal="right")
        else:
            cv = ws.cell(row=r, column=2, value=v if (v is not None and not (isinstance(v, float) and np.isnan(v))) else "N/A")
            if not isinstance(cv.value, str):
                cv.number_format = fmt
                cv.fill = PatternFill("solid", start_color=_LINK_BG)
            cv.alignment = Alignment(horizontal="right")
            cv.border = _BORDER

    _section_header(ws, 14, "Moat Tier Reference (Damodaran / Mauboussin)", 8, _GOLD)
    headers = ["Moat", "EVA Spread Threshold", "n_positive Threshold",
               "ρ (decay)", "Phase 2 Years", "Approx. half-life (yrs)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=16, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    tiers = [
        ("Wide moat",   "> 15%",  "≥ 15", 0.90, 15),
        ("Narrow moat", "> 8%",   "≥ 12", 0.83, 12),
        ("Some moat",   "> 3%",   "≥ 8",  0.75,  8),
        ("No moat",     "≤ 3%",   "< 8",  0.60,  5),
    ]
    cur_moat = val.get("moat_label", "N/A")
    for i, (lbl, sp, npos, rho_v, yr) in enumerate(tiers):
        r = 17 + i
        is_current = (lbl == cur_moat)
        bg = _GOLD if is_current else _ZEBRA
        for col, v in enumerate([lbl, sp, npos, rho_v, yr,
                                 round(np.log(0.5)/np.log(rho_v), 1)], 1):
            cc = ws.cell(row=r, column=col, value=v)
            cc.font = Font(bold=is_current)
            cc.fill = PatternFill("solid", start_color=bg)
            cc.alignment = Alignment(horizontal="center")
            cc.border = _BORDER
            if isinstance(v, float):
                cc.number_format = _FMT_DEC

    ws.merge_cells("A22:H22")
    ws.cell(row=22, column=1,
            value="  ★ 굵은 행이 현재 종목의 등급. ρ → Phase 2 기간 매핑은 EVA spread 의 지속성에 비례.").font = Font(italic=True, size=9, color="595959")

    ws.column_dimensions["A"].width = 36
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18


def _build_beta(wb, model, ticker):
    """10Y monthly returns + SLOPE/INTERCEPT/RSQ + Blume."""
    br = wb.create_sheet("Beta_Regression")
    br.merge_cells("A1:F1")
    t = br["A1"]
    t.value = f"{ticker} — Beta Estimation: 10-Year Monthly Returns vs S&P 500"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    br.row_dimensions[1].height = 32

    for c, h in enumerate(["Date", f"{ticker} Price", "S&P 500",
                            f"{ticker} Return", "S&P 500 Return"], 1):
        cell = br.cell(row=3, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    prices = getattr(model, "_monthly_prices", pd.DataFrame())
    if prices is None or prices.empty or len(prices) < 12:
        br.cell(row=4, column=1, value="(monthly price data unavailable)").font = Font(italic=True, color="595959")
        br.column_dimensions["A"].width = 36
        for col in "BCDE": br.column_dimensions[col].width = 14
        return None, None

    N = len(prices)
    for i, (date, row) in enumerate(prices.iterrows()):
        r = 4 + i
        br.cell(row=r, column=1, value=date.date()).number_format = "yyyy-mm"
        br.cell(row=r, column=2, value=float(row[ticker])).number_format = "0.00"
        br.cell(row=r, column=3, value=float(row["SP500"])).number_format = "0.00"
        if i > 0:
            br.cell(row=r, column=4, value=f"=B{r}/B{r-1}-1").number_format = _FMT_PCT3
            br.cell(row=r, column=5, value=f"=C{r}/C{r-1}-1").number_format = _FMT_PCT3
        for c in range(1, 6):
            br.cell(row=r, column=c).border = _BORDER

    RET_START = 5
    RET_END   = 4 + N - 1
    RES_HDR   = N + 5
    BETA_R    = RES_HDR + 1
    ALPHA_R   = RES_HDR + 2
    BLUME_R   = RES_HDR + 14

    _section_header(br, RES_HDR, "Regression Statistics — OLS Beta Estimation", 5, _TEAL)
    for offset, label, formula, fmt, hl in [
        (1, "Beta (slope)",                       f"=SLOPE(D{RET_START}:D{RET_END},E{RET_START}:E{RET_END})",     "0.0000", True),
        (2, "Intercept (alpha, monthly)",         f"=INTERCEPT(D{RET_START}:D{RET_END},E{RET_START}:E{RET_END})", "0.000%", False),
        (3, "Annualized alpha (×12)",             f"=B{ALPHA_R}*12",                                              "0.00%",  False),
        (4, "R² (coefficient of determination)",  f"=RSQ(D{RET_START}:D{RET_END},E{RET_START}:E{RET_END})",       "0.000",  False),
        (5, "Number of observations",             f"=COUNT(D{RET_START}:D{RET_END})",                             "0",      False),
    ]:
        c1 = br.cell(row=RES_HDR + offset, column=1, value=label)
        c1.font = Font(name="Calibri", size=10, bold=hl)
        c1.alignment = Alignment(horizontal="left")
        c1.border = _BORDER
        c2 = br.cell(row=RES_HDR + offset, column=2, value=formula)
        _set_computed(c2, formula, fmt, highlight=hl)

    _section_header(br, RES_HDR + 7, "Blume Adjustment — Mean-Reversion of Beta", 5, _TEAL)
    br.merge_cells(f"A{RES_HDR + 8}:E{RES_HDR + 8}")
    note = br.cell(row=RES_HDR + 8, column=1,
                   value="  Beta_Blume = w_raw × Beta_raw + w_market × 1.0  (industry standard: 2/3, 1/3)")
    note.font = Font(italic=True, size=10, color="595959")
    _set_label(br.cell(row=RES_HDR + 10, column=1), "Weight on raw Beta (w_raw)")
    _set_input(br.cell(row=RES_HDR + 10, column=2), _BLUME_W_RAW, _FMT_DEC)
    _set_label(br.cell(row=RES_HDR + 11, column=1), "Weight on market Beta = 1 (w_market)")
    _set_input(br.cell(row=RES_HDR + 11, column=2), _BLUME_W_MARKET, _FMT_DEC)
    bl = br.cell(row=BLUME_R, column=1, value="Beta (Blume-adjusted) ★")
    bl.font = Font(bold=True); bl.fill = PatternFill("solid", start_color=_GOLD); bl.border = _BORDER
    bv = br.cell(row=BLUME_R, column=2, value=f"=B{RES_HDR + 10}*B{BETA_R}+B{RES_HDR + 11}*1")
    bv.number_format = "0.0000"
    bv.fill = PatternFill("solid", start_color=_NAVY)
    bv.font = Font(size=11, bold=True, color="FFFFFF")
    bv.alignment = Alignment(horizontal="right"); bv.border = _BORDER

    br.column_dimensions["A"].width = 36
    for col in "BCDE": br.column_dimensions[col].width = 14
    br.freeze_panes = "A4"
    return BETA_R, BLUME_R


def _build_coc(wb, model, ticker, BETA_R, BLUME_R):
    cc = wb.create_sheet("Cost_of_Capital")
    cc.merge_cells("A1:F1")
    t = cc["A1"]
    t.value = f"{ticker} — Cost of Capital: Re (CAPM) + Rd + WACC"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    cc.row_dimensions[1].height = 32

    val = model.valuation
    inc = model._inc; bs = model._bs

    # WACC chain inputs
    interest_ttm = 0.0
    if "interestExpense" in inc.columns:
        interest_ttm = float(pd.to_numeric(inc["interestExpense"], errors="coerce").abs().tail(4).sum())
    total_debt = 0.0
    if "totalDebt" in bs.columns:
        s = pd.to_numeric(bs["totalDebt"], errors="coerce").dropna()
        total_debt = float(s.iloc[-1]) if len(s) else 0.0
    cash = 0.0
    if "cashAndCashEquivalents" in bs.columns:
        s = pd.to_numeric(bs["cashAndCashEquivalents"], errors="coerce").dropna()
        cash = float(s.iloc[-1]) if len(s) else 0.0

    # Shares: profile.mktCap/price 우선, fallback model.valuation["shares"]
    profile = getattr(model, "_profile", {}) or {}
    shares_profile = 0.0
    if profile.get("mktCap") and profile.get("price"):
        try:
            shares_profile = float(profile["mktCap"]) / float(profile["price"])
        except Exception:
            shares_profile = 0.0
    shares_model = float(val.get("shares") or 0.0) if not (isinstance(val.get("shares"), float) and np.isnan(val.get("shares"))) else 0.0
    shares_use   = shares_profile if shares_profile > 0 else shares_model

    current_price = float(val.get("current_price") or 0.0)
    if (not current_price) and profile.get("price"):
        try: current_price = float(profile["price"])
        except Exception: pass

    tax = float(val.get("wacc")) if False else float(model.estimate_tax_rate())
    wacc_db = float(val.get("wacc", np.nan)) if val else np.nan

    # Section A — Re
    _section_header(cc, 3, "Section A — Cost of Equity (Re) via CAPM", 6, _TEAL)
    cc.merge_cells("A4:F4")
    cc.cell(row=4, column=1, value="  Formula:  Re = Rf + Beta × (Rm − Rf)").font = Font(italic=True, size=10, color="595959")

    beta_raw_ref   = f"=Beta_Regression!B{BETA_R}"   if BETA_R   else 1.0
    beta_blume_ref = f"=Beta_Regression!B{BLUME_R}"  if BLUME_R  else 1.0

    spec = [
        (6,  "Risk-free rate (Rf, 10Y Treasury)",      _RF_DEFAULT,         _FMT_PCT, "input"),
        (7,  "Expected market return (Rm)",            _RM_DEFAULT,         _FMT_PCT, "input"),
        (8,  "  Market Risk Premium (MRP) = Rm − Rf",  "=B7-B6",            _FMT_PCT, "comp"),
        (10, "Beta (raw, from Beta_Regression)",       beta_raw_ref,        "0.0000", "comp"),
        (11, "Beta (Blume-adjusted)",                  beta_blume_ref,      "0.0000", "comp"),
        (13, "Re (using raw Beta)",                    "=B6+B10*B8",        _FMT_PCT, "comp"),
        (14, "Re (using Blume Beta) ★",                "=B6+B11*B8",        _FMT_PCT, "highlight"),
    ]
    for r, label, val_, fmt, kind in spec:
        lc = cc.cell(row=r, column=1)
        _set_label(lc, label, bold=(kind == "highlight"))
        if kind == "highlight":
            lc.fill = PatternFill("solid", start_color=_GOLD)
        c = cc.cell(row=r, column=2)
        if kind == "input":
            _set_input(c, val_, fmt)
        elif kind == "highlight":
            _set_computed(c, val_, fmt, highlight=True)
        else:
            _set_computed(c, val_, fmt)

    # Section B — Rd
    _section_header(cc, 17, "Section B — Cost of Debt (Rd)", 6, _TEAL)
    cc.merge_cells("A18:F18")
    cc.cell(row=18, column=1, value="  Rd_pretax = Interest Expense / Total Debt  →  Rd_aftertax = Rd_pretax × (1 − T)").font = Font(italic=True, size=10, color="595959")
    rd_spec = [
        (20, "Interest Expense (TTM, last 4Q)",     interest_ttm,                _FMT_INT, "input"),
        (21, "Total Debt (latest quarter)",         total_debt,                  _FMT_INT, "input"),
        (23, "  Rd pre-tax",                        "=IFERROR(B20/B21,0)",       _FMT_PCT, "comp"),
        (24, "Effective tax rate",                  tax,                         _FMT_PCT, "input"),
        (26, "Rd after-tax = Rd_pretax × (1 − T) ★","=IFERROR(B23*(1-B24),0)",   _FMT_PCT, "highlight"),
    ]
    for r, label, val_, fmt, kind in rd_spec:
        lc = cc.cell(row=r, column=1)
        _set_label(lc, label, bold=(kind == "highlight"))
        if kind == "highlight":
            lc.fill = PatternFill("solid", start_color=_GOLD)
        c = cc.cell(row=r, column=2)
        if kind == "input":     _set_input(c, val_, fmt)
        elif kind == "highlight": _set_computed(c, val_, fmt, highlight=True)
        else:                   _set_computed(c, val_, fmt)

    # Section C — Capital Structure
    _section_header(cc, 29, "Section C — Capital Structure (Market Values)", 6, _TEAL)
    cs_spec = [
        (31, "Current stock price",                 current_price,        _FMT_PRICE, "input"),
        (32, "Shares outstanding",                  shares_use,           _FMT_INT,   "input"),
        (33, "  Market Cap (E) = Price × Shares",   "=B31*B32",           _FMT_INT,   "comp"),
        (34, "  Total Debt (D, linked from B21)",   "=B21",               _FMT_INT,   "comp"),
        (35, "  Total Capital (V) = E + D",         "=B33+B34",           _FMT_INT,   "comp"),
        (37, "  E/V (equity weight)",               "=IFERROR(B33/B35,0)",_FMT_PCT,   "comp"),
        (38, "  D/V (debt weight)",                 "=IFERROR(B34/B35,0)",_FMT_PCT,   "comp"),
        (41, "Cash & equivalents",                  cash,                 _FMT_INT,   "input"),
        (40, "  Net Debt = Total Debt − Cash",      "=B34-B41",           _FMT_INT,   "comp"),
    ]
    for r, label, val_, fmt, kind in cs_spec:
        _set_label(cc.cell(row=r, column=1), label)
        c = cc.cell(row=r, column=2)
        if kind == "input": _set_input(c, val_, fmt)
        else:               _set_computed(c, val_, fmt)

    # Section D — WACC
    _section_header(cc, 44, "Section D — WACC", 6, _TEAL)
    cc.merge_cells("A45:F45")
    cc.cell(row=45, column=1, value="  WACC = (E/V) × Re + (D/V) × Rd × (1 − T)").font = Font(italic=True, size=10, color="595959")
    for r, label, val_, fmt, kind in [
        (47, "  Equity contribution: (E/V) × Re_Blume",  "=B37*B14", _FMT_PCT, "comp"),
        (48, "  Debt contribution: (D/V) × Rd_aftertax", "=B38*B26", _FMT_PCT, "comp"),
        (50, "WACC ★",                                    "=B47+B48", _FMT_PCT, "final"),
    ]:
        lc = cc.cell(row=r, column=1)
        _set_label(lc, label, bold=(kind == "final"))
        if kind == "final":
            lc.fill = PatternFill("solid", start_color=_GOLD)
        c = cc.cell(row=r, column=2)
        if kind == "final": _set_computed(c, val_, fmt, final=True)
        else:               _set_computed(c, val_, fmt)

    # Section E — DB-stored WACC (sanity check)
    _section_header(cc, 53, "Section E — Reference: WACC (DCFModel batch 결과)", 6, _GOLD)
    _set_label(cc.cell(row=55, column=1), "WACC (model.valuation['wacc'])")
    if not np.isnan(wacc_db):
        wc = cc.cell(row=55, column=2, value=float(wacc_db))
        wc.number_format = _FMT_PCT
        wc.fill = PatternFill("solid", start_color=_PHASE1_BG)
        wc.font = Font(bold=True); wc.alignment = Alignment(horizontal="right"); wc.border = _BORDER
    else:
        cc.cell(row=55, column=2, value="N/A").alignment = Alignment(horizontal="right")
    _set_label(cc.cell(row=56, column=1), "  Δ (Excel WACC − model WACC)")
    cc.cell(row=56, column=2, value="=B50-B55").number_format = _FMT_PCT
    cc.cell(row=56, column=2).font = Font(italic=True, color="595959")
    cc.cell(row=56, column=2).border = _BORDER

    cc.merge_cells("A58:F58")
    cc.cell(row=58, column=1, value="※ Excel WACC 와 model WACC 차이는 Rf/Rm 가정값 + Beta vintage + Rd 산출 차이에서 발생.").font = Font(italic=True, size=9, color="595959")

    cc.column_dimensions["A"].width = 52
    cc.column_dimensions["B"].width = 22
    for col in "CDEF": cc.column_dimensions[col].width = 14


def _build_dcf(wb, model, ticker):
    dv = wb.create_sheet("DCF_Valuation")
    dv.merge_cells("A1:F1")
    t = dv["A1"]
    t.value = f"{ticker} — 3-Stage DCF: WACC-Discounted FCFF → Target Price"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    dv.row_dimensions[1].height = 32

    _section_header(dv, 3, "Inputs (linked from Cost_of_Capital and Phase2_LongTerm)", 6, _TEAL)
    for r, label, val, fmt in [
        (5, "WACC",                                    "=Cost_of_Capital!B50", _FMT_PCT),
        (6, "Terminal growth (g_term)",                "=Phase2_LongTerm!B6",  _FMT_PCT),
        (7, "Net Debt",                                "=Cost_of_Capital!B40", _FMT_INT),
        (8, "Shares Outstanding",                      "=Cost_of_Capital!B32", _FMT_INT),
        (9, "Current stock price",                     "=Cost_of_Capital!B31", _FMT_PRICE),
    ]:
        _set_label(dv.cell(row=r, column=1), label)
        c = dv.cell(row=r, column=2, value=val)
        _set_computed(c, val, fmt, highlight=(r == 5))

    _section_header(dv, 12, "Year-by-Year FCFF Discount (Year 1-17)", 6, _TEAL)
    for c, h in enumerate(["Year","Cal. Year","Phase","FCFF (annual)","Discount Factor","PV of FCFF"], 1):
        cell = dv.cell(row=13, column=c, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    try:
        BASE_YEAR = int(str(model.result_df["quarter"].iloc[0])[:4])
    except Exception:
        BASE_YEAR = datetime.now().year

    for yr in range(1, 18):
        r = 13 + yr
        dv.cell(row=r, column=1, value=yr).alignment = Alignment(horizontal="center")
        dv.cell(row=r, column=2, value=BASE_YEAR - 1 + yr).alignment = Alignment(horizontal="center")
        dv.cell(row=r, column=3, value=f"=Phase2_LongTerm!C{r}").alignment = Alignment(horizontal="center")
        c4 = dv.cell(row=r, column=4, value=f"=Phase2_LongTerm!E{r}"); c4.number_format = _FMT_INT
        c5 = dv.cell(row=r, column=5, value=f"=1/(1+$B$5)^A{r}");      c5.number_format = "0.0000"
        c6 = dv.cell(row=r, column=6, value=f"=D{r}*E{r}");              c6.number_format = _FMT_INT
        bg = "E8F4F8" if yr <= 2 else "FFF4D6"
        for cc_idx in range(1, 7):
            cc = dv.cell(row=r, column=cc_idx)
            cc.border = _BORDER
            cc.fill = PatternFill("solid", start_color=bg)

    dv.merge_cells("A32:E32")
    _set_label(dv.cell(row=32, column=1), "  Σ PV of FCFF (Year 1-17)", bold=True)
    sumc = dv.cell(row=32, column=6, value="=SUM(F14:F30)")
    sumc.number_format = _FMT_INT
    sumc.fill = PatternFill("solid", start_color=_LINK_BG)
    sumc.font = Font(size=11, bold=True)
    sumc.alignment = Alignment(horizontal="right"); sumc.border = _BORDER

    _section_header(dv, 35, "Terminal Value (Phase 3, Gordon Growth Model)", 6, _TEAL)
    dv.merge_cells("A36:F36")
    dv.cell(row=36, column=1, value="  TV = FCFF_17 × (1 + g) / (WACC − g)  ;  PV = TV / (1+WACC)^17").font = Font(italic=True, size=10, color="595959")

    for r, label, val, fmt, hl in [
        (38, "FCFF at end of Year 17",                       "=D30",                  _FMT_INT, False),
        (39, "  Terminal Value = FCFF_17 × (1+g)/(WACC−g)",  "=B38*(1+B6)/(B5-B6)",   _FMT_INT, False),
        (40, "  Discount factor for Year 17",                "=1/(1+B5)^17",          "0.0000", False),
        (41, "PV of Terminal Value ★",                       "=B39*B40",              _FMT_INT, True),
    ]:
        lc = dv.cell(row=r, column=1)
        _set_label(lc, label, bold=hl)
        if hl: lc.fill = PatternFill("solid", start_color=_GOLD)
        c = dv.cell(row=r, column=2, value=val)
        _set_computed(c, val, fmt, highlight=hl)

    _section_header(dv, 44, "Enterprise Value → Equity Value → Target Price", 6, _GOLD)
    for r, label, val, fmt, kind in [
        (46, "Enterprise Value (EV)",                "=F32+B41",                _FMT_INT,  "highlight"),
        (47, "  Equity Value = EV − Net Debt",       "=B46-B7",                 _FMT_INT,  "comp"),
        (49, "Target Price ★★★",                    "=IFERROR(B47/B8, 0)",     _FMT_PRICE,"final"),
        (50, "  Current Price",                      "=B9",                     _FMT_PRICE,"comp"),
        (51, "  Upside / (Downside)",                "=IFERROR((B49-B50)/B50,0)",_FMT_PCT, "highlight"),
        (53, "(Implied Market Cap)",                 "=B49*B8",                 _FMT_INT,  "comp"),
    ]:
        lc = dv.cell(row=r, column=1)
        _set_label(lc, label, bold=(kind in ("highlight", "final")))
        if kind == "final":
            lc.fill = PatternFill("solid", start_color=_GOLD)
        c = dv.cell(row=r, column=2, value=val)
        if kind == "final":
            _set_computed(c, val, fmt, final=True)
            c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        else:
            _set_computed(c, val, fmt, highlight=(kind == "highlight"))

    # Reference (model 의 target_price)
    _section_header(dv, 56, "Reference: Target Price (DCFModel batch 결과)", 6, _GOLD)
    val_dict = model.valuation
    _set_label(dv.cell(row=58, column=1), "TP (model.valuation['target_price'])")
    tp_db = val_dict.get("target_price", np.nan)
    if tp_db is not None and not (isinstance(tp_db, float) and np.isnan(tp_db)):
        tp_c = dv.cell(row=58, column=2, value=float(tp_db))
        tp_c.number_format = _FMT_PRICE
        tp_c.fill = PatternFill("solid", start_color=_PHASE1_BG)
        tp_c.font = Font(bold=True); tp_c.alignment = Alignment(horizontal="right"); tp_c.border = _BORDER
    else:
        dv.cell(row=58, column=2, value="N/A").alignment = Alignment(horizontal="right")
    _set_label(dv.cell(row=59, column=1), "  Δ (Excel TP − model TP)")
    dv.cell(row=59, column=2, value="=B49-B58").number_format = _FMT_PRICE
    dv.cell(row=59, column=2).font = Font(italic=True, color="595959")
    dv.cell(row=59, column=2).border = _BORDER

    dv.column_dimensions["A"].width = 38
    dv.column_dimensions["B"].width = 24
    for col in "CDEF": dv.column_dimensions[col].width = 18
    dv.freeze_panes = "A14"


def _build_sens(wb, ticker):
    sn = wb.create_sheet("Sensitivity_TP")
    sn.merge_cells("A1:H1")
    t = sn["A1"]
    t.value = f"{ticker} — Target Price Sensitivity Analysis"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center")
    sn.row_dimensions[1].height = 32

    _section_header(sn, 3, "Reference Values (linked)", 8, _TEAL)
    for r, label, formula, fmt in [
        (5,  "WACC (base)",                 "=Cost_of_Capital!B50", _FMT_PCT),
        (6,  "Re_Blume (base)",             "=Cost_of_Capital!B14", _FMT_PCT),
        (7,  "Beta_Blume (base)",           "=Cost_of_Capital!B11", "0.0000"),
        (8,  "Risk-free rate (Rf)",         "=Cost_of_Capital!B6",  _FMT_PCT),
        (9,  "Expected market return (Rm)", "=Cost_of_Capital!B7",  _FMT_PCT),
        (10, "E/V",                         "=Cost_of_Capital!B37", _FMT_PCT),
        (11, "D/V",                         "=Cost_of_Capital!B38", _FMT_PCT),
        (12, "Rd after-tax",                "=Cost_of_Capital!B26", _FMT_PCT),
        (13, "FCFF Year 17",                "=Phase2_LongTerm!E30", _FMT_INT),
        (14, "Net Debt",                    "=Cost_of_Capital!B40", _FMT_INT),
        (15, "Shares Outstanding",          "=Cost_of_Capital!B32", _FMT_INT),
    ]:
        _set_label(sn.cell(row=r, column=1), label)
        c = sn.cell(row=r, column=2, value=formula)
        _set_computed(c, formula, fmt)

    g_grid = [0.02, 0.03, 0.04, 0.05, 0.06]
    sumprod_f = 'SUMPRODUCT(Phase2_LongTerm!$E$14:$E$30/(1+{w})^ROW(INDIRECT("1:17")))'

    # ── Grid 1: WACC × g_term ─────────────────────────────────
    _section_header(sn, 18, "Grid 1: TP as function of WACC × g_term", 8, _GOLD)
    for j, g in enumerate(g_grid):
        c = sn.cell(row=20, column=3 + j, value=g); c.number_format = _FMT_PCT
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", start_color=_NAVY)
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
    for i, off in enumerate([-0.02, -0.01, 0.00, 0.01, 0.02]):
        r = 21 + i
        formula_row = "=$B$5" if off == 0 else f"=$B$5{off:+}"
        c = sn.cell(row=r, column=1, value=formula_row); c.number_format = _FMT_PCT
        c.font = Font(bold=(off == 0), color="FFFFFF")
        c.fill = PatternFill("solid", start_color=_NAVY if off == 0 else "455A64")
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
        for j in range(5):
            col = 3 + j
            wacc_ref = f"$A{r}"; g_ref = f"{get_column_letter(col)}$20"
            sp = sumprod_f.format(w=wacc_ref)
            formula = (f"=({sp}+$B$13*(1+{g_ref})/({wacc_ref}-{g_ref})/(1+{wacc_ref})^17"
                       f"-$B$14)/$B$15")
            cell = sn.cell(row=r, column=col, value=formula)
            cell.number_format = _FMT_PRICE
            cell.alignment = Alignment(horizontal="right"); cell.border = _BORDER
            cell.fill = PatternFill("solid", start_color=_GOLD if (i == 2 and j == 2) else "FFFFFF")

    # ── Grid 2: Beta × g_term (WACC 재계산) ───────────────────
    _section_header(sn, 30, "Grid 2: TP as function of Beta × g_term (WACC recomputed)", 8, _GOLD)
    for j, g in enumerate(g_grid):
        c = sn.cell(row=33, column=3 + j, value=g); c.number_format = _FMT_PCT
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", start_color=_NAVY)
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
    for i, off in enumerate([-0.25, -0.10, 0.00, 0.10, 0.25]):
        r = 34 + i
        formula_row = "=$B$7" if off == 0 else f"=$B$7{off:+}"
        c = sn.cell(row=r, column=1, value=formula_row); c.number_format = "0.0000"
        c.font = Font(bold=(off == 0), color="FFFFFF")
        c.fill = PatternFill("solid", start_color=_NAVY if off == 0 else "455A64")
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
        for j in range(5):
            col = 3 + j
            beta_ref = f"$A{r}"; g_ref = f"{get_column_letter(col)}$33"
            re_local   = f"($B$8+{beta_ref}*($B$9-$B$8))"
            wacc_local = f"($B$10*{re_local}+$B$11*$B$12)"
            sp = sumprod_f.format(w=wacc_local)
            formula = (f"=({sp}+$B$13*(1+{g_ref})/({wacc_local}-{g_ref})/(1+{wacc_local})^17"
                       f"-$B$14)/$B$15")
            cell = sn.cell(row=r, column=col, value=formula)
            cell.number_format = _FMT_PRICE
            cell.alignment = Alignment(horizontal="right"); cell.border = _BORDER
            cell.fill = PatternFill("solid", start_color=_GOLD if (i == 2 and j == 2) else "FFFFFF")

    # ── Grid 3: Re × g_term ───────────────────────────────────
    _section_header(sn, 43, "Grid 3: TP as function of Re × g_term", 8, _GOLD)
    sn.merge_cells("A44:H44")
    sn.cell(row=44, column=1, value="  각 셀이 Re 를 직접 사용 → WACC 재계산. Rd 는 고정.").font = Font(italic=True, size=9, color="595959")
    for j, g in enumerate(g_grid):
        c = sn.cell(row=46, column=3 + j, value=g); c.number_format = _FMT_PCT
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", start_color=_NAVY)
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
    for i, off in enumerate([-0.02, -0.01, 0.00, 0.01, 0.02]):
        r = 47 + i
        formula_row = "=$B$6" if off == 0 else f"=$B$6{off:+}"
        c = sn.cell(row=r, column=1, value=formula_row); c.number_format = _FMT_PCT
        c.font = Font(bold=(off == 0), color="FFFFFF")
        c.fill = PatternFill("solid", start_color=_NAVY if off == 0 else "455A64")
        c.alignment = Alignment(horizontal="center"); c.border = _BORDER
        for j in range(5):
            col = 3 + j
            re_ref = f"$A{r}"; g_ref = f"{get_column_letter(col)}$46"
            wacc_local = f"($B$10*{re_ref}+$B$11*$B$12)"
            sp = sumprod_f.format(w=wacc_local)
            formula = (f"=({sp}+$B$13*(1+{g_ref})/({wacc_local}-{g_ref})/(1+{wacc_local})^17"
                       f"-$B$14)/$B$15")
            cell = sn.cell(row=r, column=col, value=formula)
            cell.number_format = _FMT_PRICE
            cell.alignment = Alignment(horizontal="right"); cell.border = _BORDER
            cell.fill = PatternFill("solid", start_color=_GOLD if (i == 2 and j == 2) else "FFFFFF")

    sn.column_dimensions["A"].width = 32
    for col in "BCDEFGH": sn.column_dimensions[col].width = 14


def _build_notes(wb, model, ticker):
    notes = wb.create_sheet("Notes")
    eva_cache = getattr(model, "_eva_cache", {}) or {}
    g0_method = eva_cache.get("g0_method", "n/a")
    notes_text = [
        (f"{ticker} — V10 Format Excel from FCFF_DCF_Valuation v9 (run: {datetime.now():%Y-%m-%d %H:%M})",
         True, 14, _NAVY, "FFFFFF"),
        ("", False, 11, None, None),
        ("Source: Financial Modeling Prep (FMP) + DCFModel.run() in-memory results", False, 11, None, None),
        ("", False, 11, None, None),
        ("Definitions", True, 12, _TEAL, "FFFFFF"),
        ("  Sales       = revenue (Income Statement)", False, 11, None, None),
        ("  EBIT        = operatingIncome", False, 11, None, None),
        ("  OPM         = EBIT / Sales", False, 11, None, None),
        ("  NOPAT       = EBIT × (1 − Tax rate)", False, 11, None, None),
        ("  D&A         = depreciationAndAmortization (Cash Flow Statement)", False, 11, None, None),
        ("  CapEx       = |capitalExpenditure| (Cash Flow Statement)", False, 11, None, None),
        ("  NWC         = (Current Assets − Cash) − (Current Liab. − Short-term Debt)", False, 11, None, None),
        ("  ΔNWC        = NWC(t) − NWC(t−1)", False, 11, None, None),
        ("  FCFF        = NOPAT + D&A − CapEx − ΔNWC", False, 11, None, None),
        ("", False, 11, None, None),
        ("Phase Structure", True, 12, _TEAL, "FFFFFF"),
        ("  Phase 1 (Year 1-2):  DB forecast 8Q (Ensemble→SARIMA→ETS→Theta) → 분기 합산 연간화", False, 11, None, None),
        ("  Phase 2 (Year 3+):   AR(1) decay g(t) = g_term + (g(t-1) − g_term) × ρ", False, 11, None, None),
        ("                       기간(n_years) 은 EVA spread → Moat 등급에서 결정", False, 11, None, None),
        ("                       ρ 는 EVA 기반 + OLS 추정 가중평균 (각 50%, clip 0.40~0.92)", False, 11, None, None),
        ("  Phase 3 (Year n+):   Gordon Growth — TV = FCFF × (1+g_term)/(WACC − g_term)", False, 11, None, None),
        ("", False, 11, None, None),
        ("g₀ Patch (v9.1) — 음수 성장률 처리", True, 12, _GOLD, "000000"),
        (f"  현재 ticker 의 g₀ method: {g0_method}", False, 11, None, None),
        ("  Method 1: 8Q CAGR (분기별 시계열 기반)", False, 11, None, None),
        ("  Method 2: 12Q rolling TTM vs TTM-2 (사이클 평활화)", False, 11, None, None),
        ("  Method 3: Phase 1 yr1/yr2 비교 (forecast 기반)", False, 11, None, None),
        ("  Method 4: G_TERM (4%) 중립 fallback", False, 11, None, None),
        ("  하한: -20% (사이클릭 침체 허용), 상한: +80%", False, 11, None, None),
        ("", False, 11, None, None),
        ("Yellow cells (입력) — 변경 시 자동 재계산", True, 11, _GOLD, "000000"),
        ("  Main sheet: M11:M15 (가정), N4:N7 (분기별 성장률)", False, 11, None, None),
        ("  Phase2_LongTerm: B4:B7 (ρ, g₀, g_term, 기간)", False, 11, None, None),
        ("  Cost_of_Capital: B6,B7 (Rf/Rm), B20,B21,B41 (debt/interest/cash), B24 (tax)", False, 11, None, None),
        ("", False, 11, None, None),
        ("Caveats", True, 12, _TEAL, "FFFFFF"),
        ("  • 마진 일정 가정 (constant margin) — 실제 DCF 는 단계적 margin 수렴 권고 (Damodaran)", False, 11, None, None),
        ("  • Phase 2 AR(1) 음수 g₀ 시 decay 가 GDP 성장률(4%) 로 자연 회복 — 평균 회귀 (Fama-French 1995)", False, 11, None, None),
        ("  • Cost_of_Capital, DCF_Valuation 의 'Reference: model' 행에서 batch 결과와 Δ 비교 가능", False, 11, None, None),
        ("  • FMP Starter plan 한계로 일부 BS 필드 (cash 등) 가 0 일 수 있음 → 노란 셀 직접 입력 권장", False, 11, None, None),
    ]
    for i, (txt, bold, size, fc, font_c) in enumerate(notes_text, 1):
        c = notes.cell(row=i, column=1, value=txt)
        c.font = Font(name="Calibri", size=size, bold=bold,
                      color=font_c if font_c else "000000")
        if fc: c.fill = PatternFill("solid", start_color=fc)
        c.alignment = Alignment(horizontal="left", vertical="center")
    notes.column_dimensions["A"].width = 130


# ── Main entry: model.export_v10_excel(out_dir) ─────────────────

def export_v10_excel(self, out_dir=None, fetch_extras=True) -> Path:
    """
    DCFModel.run() 직후 호출.  9-sheet workbook 생성 후 저장.
    Returns
    -------
    Path : 저장된 .xlsx 파일 경로
    """
    if self.valuation is None or self.result_df is None:
        raise RuntimeError(
            f"[{self.ticker}] export_v10_excel 호출 전에 model.run() 을 먼저 실행하세요."
        )

    # 출력 경로
    if out_dir is None:
        out_dir = Path(r"C:/reports") if os.name == "nt" else Path.home() / "reports"
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{self.ticker}_FCFF_v10format.xlsx"

    # 추가 fetch
    if fetch_extras:
        self._fetch_v10_extras(years=10)

    if self.verbose:
        log(self.ticker, f"V10 Excel 빌드 시작 → {out_file}")

    # 과거 40Q 구성요소 + 예측 8Q
    hist_df = _build_hist_components(self, n_qtrs=40)
    if hist_df.empty:
        raise RuntimeError(f"[{self.ticker}] 과거 FCFF 구성요소 생성 실패")
    fc_df = self.result_df.copy()
    if "quarter" not in fc_df.columns:
        fc_df["quarter"] = fc_df["date"].apply(lambda d: f"{d.year}Q{d.quarter}")

    wb = Workbook()
    _build_main(self, wb, hist_df, fc_df, self.ticker)            # 1
    _build_verification(wb, hist_df, self.ticker)                 # 2
    _build_phase2(wb, self, self.ticker)                          # 3
    _build_moat(wb, self, self.ticker)                            # 4
    BETA_R, BLUME_R = _build_beta(wb, self, self.ticker)          # 5
    _build_coc(wb, self, self.ticker, BETA_R, BLUME_R)            # 6
    _build_dcf(wb, self, self.ticker)                             # 7
    _build_sens(wb, self.ticker)                                  # 8
    _build_notes(wb, self, self.ticker)                           # 9

    wb.save(out_file)
    if self.verbose:
        log(self.ticker, f"✓ V10 Excel 저장 완료 ({len(wb.sheetnames)} sheets)")

    return out_file


DCFModel.export_v10_excel = export_v10_excel


print("[OK] DCFModel patched:")
print("     ① _estimate_phase2_growth (v9.1, negative g₀ handling)")
print("     ② _fetch_v10_extras       (10Y monthly + profile)")
print("     ③ export_v10_excel        (9-sheet workbook)")
