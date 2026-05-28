"""
dataguide_fs_loader.py
======================
DataGuide(Wisereport) 엑셀 파일을 long-format으로 변환해
MariaDB `investar.korea_fs_data_from_DG` 테이블에 적재하는 모듈.

입력 파일 구조
--------------
- FCFF_RIM_DATA.xlsx    (KS/코스피)
- FCFF_RIM_KQ_DATA.xlsx (KQ/코스닥)
- 각 파일: 4개 시트(IS, BS, CF, stock)
- 시트 내부: wide-format
    Row 9  : 코드       (A005930 등)
    Row 10 : 코드명      (삼성전자 등)
    Row 11 : 유형       (NFS-IFRS(M) 등, 참고용)
    Row 12 : 아이템코드  (M000904001 등)
    Row 13 : 아이템명    (매출액(천원) 등)
    Row 14 : 집계주기   (4Q / 일간)
    Row 15~: 날짜       | 값 | 값 | ...

출력 long-format 스키마
-----------------------
korea_fs_data_from_DG:
    date           DATE        -- 분기말 또는 주가일자
    ticker         VARCHAR(10) -- '005930' (A 제거, 6자리)
    company_name   VARCHAR(100)
    item_code      VARCHAR(20) -- 'M000904001' 등 DataGuide 원본 코드
    indicator      VARCHAR(200)-- '매출액(천원)' 등
    sj_div         VARCHAR(10) -- IS / BS / CF / stock
    market         VARCHAR(10) -- KS / KQ
    value          DOUBLE
    freq           VARCHAR(10) -- '4Q' / '일간' 등
    created_at     TIMESTAMP
    updated_at     TIMESTAMP
    PRIMARY KEY (date, ticker, item_code)

중복 방지
---------
PK = (date, ticker, item_code) 로 설정하여
동일 시점·동일 기업·동일 지표는 단 하나만 존재하도록 강제.
재실행 시 ON DUPLICATE KEY UPDATE 로 값만 최신화.
"""

from __future__ import annotations

import os
import sys
import time
import logging
from pathlib import Path
from typing import List, Dict, Iterator, Tuple, Optional

import openpyxl
import pymysql
from datetime import datetime, date

# ---------------------------------------------------------
# 로깅
# ---------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ==========================================================
# 상단 상수 (필요 시 여기만 조정)
# ==========================================================
TICKER_STRIP_A       = False   # True: 'A005930' → '005930', False: 원본 유지 ('A005930')
INDICATOR_KEEP_UNIT  = True    # True: '매출액(천원)' 보존, False: '(천원)' 제거
TABLE_NAME           = "korea_fs_data_from_DG"
BATCH_INSERT_SIZE    = 5000    # executemany 배치 크기
META_ROWS = {
    'code':        9,
    'name':        10,
    'type':        11,
    'item_code':   12,
    'item_name':   13,
    'freq':        14,
}
DATA_START_ROW       = 15      # 날짜 데이터가 시작되는 행

SHEET_TO_SJ_DIV = {
    'IS':    'IS',
    'BS':    'BS',
    'CF':    'CF',
    'stock': 'stock',
}


# ==========================================================
# 유틸
# ==========================================================
def normalize_ticker(raw: str) -> Optional[str]:
    """'A005930' → '005930'. 코드 없으면 None 반환."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if TICKER_STRIP_A and s.upper().startswith('A'):
        s = s[1:]
    return s.zfill(6) if s.isdigit() else s


def normalize_indicator(raw: str) -> Optional[str]:
    """아이템명 정규화. 필요 시 단위 괄호 제거."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if not INDICATOR_KEEP_UNIT:
        # '매출액(천원)' → '매출액'
        idx = s.find('(')
        if idx > 0:
            s = s[:idx].strip()
    return s


def normalize_date(raw) -> Optional[date]:
    """엑셀 날짜 셀 → python date."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    # 문자열이면 파싱 시도
    try:
        return datetime.strptime(str(raw)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def normalize_value(raw) -> Optional[float]:
    """값 셀 → float. 공백/None/숫자 외 모두 None."""
    if raw is None or raw == "":
        return None
    try:
        v = float(raw)
        # NaN 필터
        if v != v:
            return None
        return v
    except (TypeError, ValueError):
        return None


# ==========================================================
# 엑셀 파싱
# ==========================================================
def iter_rows_from_sheet(
    ws,
    market: str,
    sj_div: str,
) -> Iterator[Tuple]:
    """
    Wide format 시트를 순회하며 long format 튜플을 생성.

    Yields
    ------
    (date, ticker, company_name, item_code, indicator, sj_div, market, value, freq)
    """
    max_col = ws.max_column
    max_row = ws.max_row

    # 1) 헤더 로드 — 한 번에 읽기
    def _read_header_row(row_num: int) -> List:
        return list(ws.iter_rows(
            min_row=row_num, max_row=row_num,
            min_col=2, max_col=max_col,
            values_only=True
        ))[0]

    codes        = _read_header_row(META_ROWS['code'])
    names        = _read_header_row(META_ROWS['name'])
    item_codes   = _read_header_row(META_ROWS['item_code'])
    item_names   = _read_header_row(META_ROWS['item_name'])
    freqs        = _read_header_row(META_ROWS['freq'])

    n_data_cols = len(codes)
    logger.info(f"[{market}/{sj_div}] 데이터 컬럼 수: {n_data_cols}")

    # 2) 각 컬럼의 메타데이터를 미리 정규화 (매 행마다 반복 계산 방지)
    col_meta = []
    for i in range(n_data_cols):
        ticker       = normalize_ticker(codes[i])
        company_name = (str(names[i]).strip() if names[i] is not None else None)
        item_code    = (str(item_codes[i]).strip() if item_codes[i] is not None else None)
        indicator    = normalize_indicator(item_names[i])
        freq         = (str(freqs[i]).strip() if freqs[i] is not None else None)

        # 헤더가 불완전한 컬럼은 표시만 해두고 추후 skip
        valid = bool(ticker and item_code and indicator)
        col_meta.append((ticker, company_name, item_code, indicator, freq, valid))

    # 3) 날짜 행 순회
    for row in ws.iter_rows(
        min_row=DATA_START_ROW, max_row=max_row,
        min_col=1, max_col=max_col,
        values_only=True
    ):
        dt = normalize_date(row[0])
        if dt is None:
            continue

        # 각 데이터 컬럼의 값 추출
        for i in range(n_data_cols):
            ticker, company_name, item_code, indicator, freq, valid = col_meta[i]
            if not valid:
                continue

            val = normalize_value(row[i + 1])  # row[0]은 날짜
            if val is None:
                continue   # None/NaN은 저장 안 함

            yield (
                dt, ticker, company_name,
                item_code, indicator, sj_div, market,
                val, freq,
            )


def parse_excel_file(
    file_path: Path,
    market: str,
) -> Iterator[Tuple]:
    """엑셀 1개 파일 4개 시트 전체를 순회하며 long 레코드 산출."""
    logger.info(f"엑셀 로드 시작: {file_path.name} (market={market})")
    wb = openpyxl.load_workbook(
        str(file_path),
        read_only=True,   # 메모리 절약 (대용량)
        data_only=True,   # 수식 결과값만
    )
    try:
        for sheet_name, sj_div in SHEET_TO_SJ_DIV.items():
            if sheet_name not in wb.sheetnames:
                logger.warning(f"  시트 '{sheet_name}' 없음 → skip")
                continue
            ws = wb[sheet_name]
            logger.info(f"  시트 '{sheet_name}' 파싱 중... (rows={ws.max_row}, cols={ws.max_column})")
            yield from iter_rows_from_sheet(ws, market=market, sj_div=sj_div)
    finally:
        wb.close()


# ==========================================================
# DB 로더
# ==========================================================
DDL_CREATE_TABLE = f"""
CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
    date          DATE          NOT NULL,
    ticker        VARCHAR(10)   NOT NULL,
    company_name  VARCHAR(200),
    item_code     VARCHAR(20)   NOT NULL,
    indicator     VARCHAR(200),
    sj_div        VARCHAR(10),
    market        VARCHAR(10),
    value         DOUBLE,
    freq          VARCHAR(10),
    created_at    TIMESTAMP     DEFAULT CURRENT_TIMESTAMP,
    updated_at    TIMESTAMP     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
    PRIMARY KEY (date, ticker, item_code),
    INDEX idx_ticker (ticker),
    INDEX idx_date (date),
    INDEX idx_indicator (indicator),
    INDEX idx_market_sj (market, sj_div)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
  COMMENT='DataGuide wide format → long format 변환 저장소'
"""

INSERT_SQL = f"""
INSERT INTO {TABLE_NAME} (
    date, ticker, company_name, item_code, indicator,
    sj_div, market, value, freq
) VALUES (
    %s, %s, %s, %s, %s, %s, %s, %s, %s
)
ON DUPLICATE KEY UPDATE
    company_name = VALUES(company_name),
    indicator    = VALUES(indicator),
    sj_div       = VALUES(sj_div),
    market       = VALUES(market),
    value        = VALUES(value),
    freq         = VALUES(freq),
    updated_at   = CURRENT_TIMESTAMP
"""


def create_table_if_not_exists(conn):
    with conn.cursor() as cur:
        cur.execute(DDL_CREATE_TABLE)
    conn.commit()
    logger.info(f"테이블 확인/생성 완료: {TABLE_NAME}")


def drop_and_create_table(conn):
    """
    initial 모드 전용: 기존 테이블을 DROP 후 새 스키마로 재생성.

    TRUNCATE와 달리 스키마 변경(신규 컬럼 추가 등)까지 반영된다.
    기존 데이터는 모두 삭제되므로 주의.
    """
    with conn.cursor() as cur:
        cur.execute(f"DROP TABLE IF EXISTS {TABLE_NAME}")
        cur.execute(DDL_CREATE_TABLE)
    conn.commit()
    logger.warning(f"[DROP+CREATE] {TABLE_NAME} 테이블 재생성 완료 (스키마 최신화)")


def bulk_insert(conn, rows: List[Tuple]) -> int:
    """batch insert with ON DUPLICATE KEY UPDATE."""
    if not rows:
        return 0
    with conn.cursor() as cur:
        cur.executemany(INSERT_SQL, rows)
    conn.commit()
    return len(rows)


# ==========================================================
# 메인 워크플로우
# ==========================================================
def load_dataguide_to_db(
    file_configs: List[Dict],
    db_info: Dict,
    mode: str = "update",    # "initial" = DROP+CREATE 후 적재, "update" = upsert만
):
    """
    DataGuide 엑셀들을 DB에 적재.

    Parameters
    ----------
    file_configs : List[Dict]
        [
            {"path": Path("..."), "market": "KS"},
            {"path": Path("..."), "market": "KQ"},
        ]
    db_info : Dict
        pymysql 연결 정보 (host, port, user, password, database)
    mode : str
        "initial": 테이블 DROP 후 새 스키마로 재생성 (기존 데이터 전부 제거, 스키마 최신화)
        "update" : 기존 테이블 유지, 동일 PK만 upsert (테이블이 없으면 생성)
    """
    t_start = datetime.now()

    # 0) 파일 존재 확인
    for cfg in file_configs:
        if not cfg["path"].exists():
            raise FileNotFoundError(f"엑셀 파일이 없습니다: {cfg['path']}")

    # 1) DB 연결
    conn = pymysql.connect(
        host=db_info["host"], port=db_info["port"],
        user=db_info["user"], password=db_info["password"],
        database=db_info["database"],
        charset="utf8mb4",
        autocommit=False,
    )

    try:
        if mode == "initial":
            logger.warning("=" * 70)
            logger.warning("⚠️  MODE: INITIAL — 테이블 DROP 후 새 스키마로 재생성")
            logger.warning("=" * 70)
            drop_and_create_table(conn)
        elif mode == "update":
            logger.info("=" * 70)
            logger.info("MODE: UPDATE — 기존 데이터 유지, upsert만 실행")
            logger.info("=" * 70)
            create_table_if_not_exists(conn)
        else:
            raise ValueError(f"알 수 없는 mode: {mode} (initial/update 중 선택)")

        # 2) 파일별 적재
        grand_total = 0
        for cfg in file_configs:
            file_start = datetime.now()
            logger.info(f"\n{'='*70}\n[파일 시작] {cfg['path'].name} (market={cfg['market']})\n{'='*70}")

            buffer: List[Tuple] = []
            inserted_for_this_file = 0

            for rec in parse_excel_file(cfg["path"], cfg["market"]):
                buffer.append(rec)
                if len(buffer) >= BATCH_INSERT_SIZE:
                    inserted_for_this_file += bulk_insert(conn, buffer)
                    logger.info(f"  ... {inserted_for_this_file:,} rows upserted")
                    buffer.clear()

            # 남은 버퍼 flush
            if buffer:
                inserted_for_this_file += bulk_insert(conn, buffer)

            grand_total += inserted_for_this_file
            file_elapsed = datetime.now() - file_start
            logger.info(
                f"[파일 완료] {cfg['path'].name}: "
                f"{inserted_for_this_file:,} rows (소요 {file_elapsed})"
            )

        # 3) 최종 요약
        total_elapsed = datetime.now() - t_start
        logger.info(f"\n{'='*70}")
        logger.info(f"[전체 완료] 총 {grand_total:,} rows 처리, 소요 {total_elapsed}")
        logger.info(f"{'='*70}")

        return grand_total

    finally:
        conn.close()


# ==========================================================
# CLI 실행
# ==========================================================
if __name__ == "__main__":
    # ----- 엑셀 폴더 경로 (2대 PC 자동 감지) -----
    EXCEL_DIR_CANDIDATES = [
        # 데스크탑 (user=82108)
        Path(r"C:\Users\82108\OneDrive\INVESTMENT\한국주식\FCFF_RIM_재무데이터"),
        # 노트북 (user=Hoyoung_Park) — 실제 경로 다르면 여기에 맞춰 수정
        Path(r"C:\Users\Hoyoung_Park\OneDrive\INVESTMENT\한국주식\FCFF_RIM_재무데이터"),
    ]
    EXCEL_DIR = None
    for candidate in EXCEL_DIR_CANDIDATES:
        if candidate.exists():
            EXCEL_DIR = candidate
            break
    if EXCEL_DIR is None:
        logger.error("DataGuide 엑셀 폴더를 찾을 수 없습니다.")
        for c in EXCEL_DIR_CANDIDATES:
            logger.error(f"  시도: {c}")
        sys.exit(1)

    logger.info(f"Excel dir: {EXCEL_DIR}")

    file_configs = [
        {"path": EXCEL_DIR / "FCFF_RIM_DATA_2026_1Q.xlsx",    "market": "KS"},
        {"path": EXCEL_DIR / "FCFF_RIM_KQ_DATA_2026_1Q.xlsx", "market": "KQ"},
    ]

    # DB 연결 정보 (환경에 맞게 조정)
    try:
        from DATA.stock_invest_function import get_db_host
        db_info = {
            "host":     get_db_host(),
            "port":     3307,
            "user":     "stox7412",
            "password": "Apt106503!~",
            "database": "investar",
        }
    except ImportError:
        logger.error("stock_invest_function import 실패. 수동으로 db_info 설정 필요.")
        sys.exit(1)

    # 모드: 첫 적재는 "initial", 이후 수시 업데이트는 "update"
    mode = sys.argv[1] if len(sys.argv) > 1 else "initial"

    load_dataguide_to_db(file_configs, db_info, mode=mode)
